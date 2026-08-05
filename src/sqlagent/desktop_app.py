"""
SQLAgent Desktop — PySide6 桌面端
Qt for Python 现代 GUI
运行: python src/sqlagent/desktop_app.py
"""
import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import requests
from PySide6 import QtCore
from PySide6.QtCore import (Qt, QThread, Signal, QRect, QRegularExpression, QSize)
from PySide6.QtGui import (QAction, QColor, QFont, QFontDatabase,
                            QKeySequence, QSyntaxHighlighter,
                            QTextCharFormat, QPalette, QIcon, QAction)
from PySide6.QtWidgets import (QApplication, QCheckBox, QComboBox,
                                QDialog, QDialogButtonBox, QFileDialog,
                                QFormLayout, QFrame, QHBoxLayout, QHeaderView,
                                QLabel, QLineEdit, QMainWindow, QMenu, QMenuBar,
                                QMessageBox,
                                QPlainTextEdit, QPushButton, QSizePolicy,
                                QSpacerItem, QSplitter, QStatusBar, QStyle,
                                QTabBar, QTabWidget, QTableWidget, QTableWidgetItem,
                                QTextEdit, QTreeWidget, QTreeWidgetItem,
                                QVBoxLayout, QWidget, QListWidget,
                                QListWidgetItem, QGroupBox, QStackedWidget)

# ═══════════════════════════════════════════
# 配色 & 常量
# ═══════════════════════════════════════════
API_BASE = 'http://localhost:8000'
ICONS_DIR = Path(__file__).parent / 'static'
BG = '#2B2B2B'
PANEL = '#3C3F41'
INPUT_BG = '#3C3F41'
MUTED = '#808080'
SUCCESS_COLOR = '#6A8759'
DANGER_COLOR = '#BC3F3C'
WARNING_COLOR = '#CC7832'
ACCENT_COLOR = '#00BFA5'
GRADIENT_START = '#00BFA5'
GRADIENT_MID = '#00E676'
GRADIENT_END = '#00B0FF'

DB_CONFIG_FILE = Path(__file__).parent.parent.parent / '.db_configs.json'
HISTORY_FILE = Path.home() / '.sqlagent_history.json'

DANGER_KW = ['DROP TABLE', 'DROP DATABASE', 'TRUNCATE', 'DELETE FROM']

# ═══════════════════════════════════════════
# 暗色主题 Stylesheet
# ═══════════════════════════════════════════
DARK_QSS = """
QMainWindow, QWidget { background: #2B2B2B; color: #A9B7C6; font-family: "JetBrains Mono","Consolas","Microsoft YaHei"; font-size: 13px; }
QGroupBox { border: 1px solid rgba(255,255,255,0.06); border-radius: 4px; margin-top: 14px; padding-top: 14px; font-weight: bold; color: #86909C; }
QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 6px; }
QLineEdit, QPlainTextEdit, QTextEdit {
    background: #3C3F41; color: #A9B7C6; border: 1px solid rgba(255,255,255,0.06);
    border-radius: 4px; padding: 6px 8px; selection-background-color: #00BFA5;
}
QPlainTextEdit:focus, QTextEdit:focus, QLineEdit:focus {
    border: 1px solid #00BFA5;
}
QComboBox {
    background: #3C3F41; color: #A9B7C6; border: 1px solid rgba(255,255,255,0.06);
    border-radius: 4px; padding: 4px 8px; min-height: 20px;
}
QComboBox:hover { border-color: rgba(255,255,255,0.12); }
QComboBox::drop-down { border: none; width: 20px; }
QComboBox QAbstractItemView {
    background: #3C3F41; color: #A9B7C6; selection-background-color: #00BFA5;
    border: 1px solid rgba(255,255,255,0.06); outline: none;
}
QPushButton {
    background: #3C3F41; color: #A9B7C6; border: 1px solid rgba(255,255,255,0.08);
    border-radius: 4px; padding: 5px 14px; min-height: 26px;
}
QPushButton:hover { background: #4E5254; border-color: rgba(255,255,255,0.15); }
QPushButton:pressed { background: #00BFA5; }
QPushButton[accent="true"] { background: #00BFA5; border: none; font-weight: bold; }
QPushButton[accent="true"]:hover { background: #00897B; }
QPushButton[danger="true"] { background: #E05555; border: none; }
QPushButton:disabled { background: #3C3F41; color: #5A6270; }
QTableWidget {
    background: #3C3F41; color: #A9B7C6; gridline-color: rgba(255,255,255,0.04);
    border: none; selection-background-color: #00BFA5;
}
QTableWidget::item { padding: 2px 6px; }
QHeaderView::section {
    background: #3C3F41; color: #86909C; border: none; border-bottom: 2px solid rgba(255,255,255,0.06);
    padding: 4px 8px; font-weight: bold; font-size: 11px;
}
QTabWidget::pane { border: none; background: #2B2B2B; }
QTabBar::tab {
    background: #3C3F41; color: #86909C; border: none; padding: 6px 24px 6px 12px;
    margin-right: 2px; border-top-left-radius: 4px; border-top-right-radius: 4px;
}
QTabBar::tab:selected { background: #00BFA5; color: #A9B7C6; }
QTabBar::tab:hover:!selected { background: #4E5254; }
QSplitter::handle { background: rgba(255,255,255,0.04); }
QSplitter::handle:hover { background: #00BFA5; }
QScrollBar:vertical { background: #2B2B2B; width: 6px; }
QScrollBar::handle:vertical { background: #5A6270; border-radius: 3px; min-height: 30px; }
QScrollBar::handle:vertical:hover { background: #86909C; }
QScrollBar:horizontal { background: #2B2B2B; height: 6px; }
QScrollBar::handle:horizontal { background: #5A6270; border-radius: 3px; min-width: 30px; }
QScrollBar::add-line, QScrollBar::sub-line { height: 0; width: 0; }
QStatusBar { background: #2B2B2B; color: #86909C; border-top: 1px solid rgba(255,255,255,0.06); font-size: 11px; }
QListWidget { background: #3C3F41; color: #A9B7C6; border: none; outline: none; }
QListWidget::item { padding: 6px 10px; border-bottom: 1px solid rgba(255,255,255,0.03); }
QListWidget::item:hover { background: #4E5254; }
QListWidget::item:selected { background: #00BFA5; }
QTreeWidget { background: #3C3F41; color: #A9B7C6; border: none; outline: none; }
QTreeWidget::item { padding: 3px 4px; }
QTreeWidget::item:hover { background: #4E5254; }
QTreeWidget::item:selected { background: #00BFA5; }
QTreeWidget::branch:has-children:!has-siblings:closed,
QTreeWidget::branch:closed:has-children:has-siblings { border-image: none; }
QCheckBox { color: #86909C; spacing: 6px; }
QCheckBox::indicator { width: 14px; height: 14px; border: 1px solid rgba(255,255,255,0.15); border-radius: 2px; }
QCheckBox::indicator:checked { background: #00BFA5; border-color: #00BFA5; }
QDialog { background: #2B2B2B; }
QMenu { background: #3C3F41; color: #A9B7C6; border: 1px solid rgba(255,255,255,0.06); padding: 4px; }
QMenu::item { padding: 6px 24px; border-radius: 2px; }
QMenu::item:selected { background: #00BFA5; }
"""

# ═══════════════════════════════════════════
# SQL 语法高亮器
# ═══════════════════════════════════════════
SQL_KEYWORDS = [
    'SELECT', 'FROM', 'WHERE', 'AND', 'OR', 'NOT', 'IN', 'LIKE', 'BETWEEN',
    'JOIN', 'LEFT', 'RIGHT', 'INNER', 'OUTER', 'ON', 'AS', 'GROUP', 'BY',
    'ORDER', 'HAVING', 'LIMIT', 'OFFSET', 'UNION', 'INSERT', 'INTO',
    'VALUES', 'UPDATE', 'SET', 'DELETE', 'DROP', 'CREATE', 'ALTER', 'TABLE',
    'COUNT', 'SUM', 'AVG', 'MAX', 'MIN', 'DISTINCT', 'EXPLAIN', 'INDEX',
    'PRIMARY', 'KEY', 'FOREIGN', 'REFERENCES', 'NULL', 'DEFAULT',
    'AUTO_INCREMENT', 'CASCADE', 'INFORMATION_SCHEMA', 'TABLE_NAME',
    'COLUMN_NAME', 'DATE_SUB', 'DATE_ADD', 'NOW', 'VARCHAR', 'INT',
    'BIGINT', 'DECIMAL', 'TEXT', 'DATETIME', 'TIMESTAMP', 'IF', 'EXISTS',
    'SHOW', 'DESCRIBE', 'USE', 'TRUNCATE', 'RENAME', 'REPLACE', 'MERGE',
    'ASC', 'DESC', 'IS', 'CASE', 'WHEN', 'THEN', 'ELSE', 'END', 'ALL',
    'ANY', 'SOME', 'FULL', 'CROSS', 'NATURAL', 'USING', 'UNIQUE', 'CHECK',
    'CONSTRAINT', 'ADD', 'COLUMN', 'MODIFY', 'CHANGE', 'RENAME', 'TO',
    'GRANT', 'REVOKE', 'ON', 'SCHEMA', 'DATABASE', 'VIEW', 'PROCEDURE',
    'FUNCTION', 'TRIGGER', 'EVENT', 'CHARACTER', 'COLLATE', 'ENGINE',
    'INNODB', 'MYISAM', 'CHARSET', 'UTF8MB4',
]

class SqlHighlighter(QSyntaxHighlighter):
    """SQL 语法高亮器"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.rules: List[tuple] = []

        # 关键字 (Darcula橙)
        kw_fmt = QTextCharFormat()
        kw_fmt.setForeground(QColor('#CC7832'))
        kw_fmt.setFontWeight(QFont.Bold)
        for kw in SQL_KEYWORDS:
            pattern = QRegularExpression(
                r'\b' + kw.replace(' ', r'\s+') + r'\b',
                QRegularExpression.CaseInsensitiveOption
            )
            self.rules.append((pattern, kw_fmt))

        # 字符串 (Darcula绿)
        str_fmt = QTextCharFormat()
        str_fmt.setForeground(QColor('#6A8759'))
        self.rules.append((QRegularExpression(r"'[^']*'"), str_fmt))
        self.rules.append((QRegularExpression(r'"[^"]*"'), str_fmt))

        # 数字 (Darcula蓝)
        num_fmt = QTextCharFormat()
        num_fmt.setForeground(QColor('#6897BB'))
        self.rules.append((QRegularExpression(r'\b\d+\.?\d*\b'), num_fmt))

        # 注释 (灰色斜体)
        cmt_fmt = QTextCharFormat()
        cmt_fmt.setForeground(QColor(MUTED))
        cmt_fmt.setFontItalic(True)
        self.rules.append((QRegularExpression(r'--[^\n]*'), cmt_fmt))
        self.rules.append((QRegularExpression(r'/\*.*?\*/',
                           QRegularExpression.DotMatchesEverythingOption), cmt_fmt))

    def highlightBlock(self, text):
        for pattern, fmt in self.rules:
            it = pattern.globalMatch(text)
            while it.hasNext():
                m = it.next()
                self.setFormat(m.capturedStart(), m.capturedLength(), fmt)


# ═══════════════════════════════════════════
# 后台工作线程
# ═══════════════════════════════════════════
class ApiWorker(QThread):
    """后台 API 调用线程"""
    finished = Signal(object)

    def __init__(self, target, *args):
        super().__init__()
        self._target = target
        self._args = args

    def run(self):
        try:
            result = self._target(*self._args)
        except Exception as e:
            result = {'error': str(e), 'success': False}
        self.finished.emit(result)


# ═══════════════════════════════════════════
# 数据库配置存储
# ═══════════════════════════════════════════
_db_configs: List[Dict[str, Any]] = []

def load_db_configs():
    global _db_configs
    try:
        if DB_CONFIG_FILE.exists():
            _db_configs = json.loads(DB_CONFIG_FILE.read_text('utf-8'))
    except Exception:
        _db_configs = []

def save_db_configs():
    DB_CONFIG_FILE.write_text(json.dumps(_db_configs, ensure_ascii=False, indent=2), 'utf-8')

def load_history() -> List[Dict]:
    try:
        if HISTORY_FILE.exists():
            return json.loads(HISTORY_FILE.read_text('utf-8'))
    except Exception:
        pass
    return []

def save_history(items: List[Dict]):
    HISTORY_FILE.write_text(json.dumps(items[-200:], ensure_ascii=False, indent=2), 'utf-8')


# ═══════════════════════════════════════════
# DB 连接弹窗
# ═══════════════════════════════════════════
class DbConfigDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('数据库连接配置')
        self.setMinimumWidth(420)
        layout = QFormLayout(self)
        layout.setSpacing(8)

        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText('我的数据库')
        layout.addRow('名称:', self.name_edit)

        self.type_combo = QComboBox()
        self.type_combo.addItems(['mysql', 'postgresql', 'sqlserver', 'sqlite'])
        layout.addRow('类型:', self.type_combo)

        self.host_edit = QLineEdit('localhost')
        layout.addRow('地址:', self.host_edit)

        self.port_edit = QLineEdit('3306')
        layout.addRow('端口:', self.port_edit)

        self.user_edit = QLineEdit('root')
        layout.addRow('用户名:', self.user_edit)

        self.pass_edit = QLineEdit()
        self.pass_edit.setEchoMode(QLineEdit.Password)
        layout.addRow('密码:', self.pass_edit)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addRow(btns)

    def get_config(self) -> Dict[str, Any]:
        return {
            'id': datetime.now().strftime('%Y%m%d%H%M%S'),
            'name': self.name_edit.text(),
            'type': self.type_combo.currentText(),
            'host': self.host_edit.text(),
            'port': int(self.port_edit.text()) if self.port_edit.text().isdigit() else 3306,
            'user': self.user_edit.text(),
            'password': self.pass_edit.text(),
            'database': '',  # 不指定库，连接服务器根级别
        }


# ═══════════════════════════════════════════
# 主窗口
# ═══════════════════════════════════════════
# ═══════════════════════════════════════════
# 表设计器弹窗
# ═══════════════════════════════════════════
class TableDesignerDialog(QDialog):
    """可视化表结构编辑器"""
    def __init__(self, db_name: str, table_name: str, columns: List[Dict], parent=None):
        super().__init__(parent)
        self.db_name = db_name
        self.table_name = table_name
        self.orig_columns = columns  # 原始列信息
        self.setWindowTitle(f'设计表: {db_name}.{table_name}')
        self.setMinimumSize(700, 400)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f'<b>{self.db_name}.{self.table_name}</b> — 双击单元格编辑'))

        # 可编辑表格
        self.tbl = QTableWidget()
        self.tbl.setColumnCount(5)
        self.tbl.setHorizontalHeaderLabels(['列名', '类型', '可空', '默认值', '键(PRI/UNI/MUL)'])
        self.tbl.horizontalHeader().setStretchLastSection(True)
        self._load_columns()
        layout.addWidget(self.tbl, 1)

        # 按钮
        btn_row = QHBoxLayout()
        add_btn = QPushButton('+ 添加列')
        add_btn.clicked.connect(self._add_column)
        btn_row.addWidget(add_btn)
        del_btn = QPushButton('🗑 删除选中列')
        del_btn.clicked.connect(self._delete_column)
        btn_row.addWidget(del_btn)
        btn_row.addStretch()
        save_icon = QIcon(str(ICONS_DIR / 'diskette.png'))
        save_btn = QPushButton(save_icon, '保存修改')
        save_btn.setProperty('accent', True)
        save_btn.clicked.connect(self._save)
        btn_row.addWidget(save_btn)
        cancel_btn = QPushButton('取消')
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)
        layout.addLayout(btn_row)

    def _load_columns(self):
        self.tbl.setRowCount(len(self.orig_columns))
        for r, col in enumerate(self.orig_columns):
            items = [
                QTableWidgetItem(col.get('name', '')),
                QTableWidgetItem(col.get('type', '')),
                QTableWidgetItem('YES' if col.get('nullable') else 'NO'),
                QTableWidgetItem(str(col.get('default', '')) if col.get('default') else ''),
                QTableWidgetItem(col.get('key', '')),
            ]
            for c, item in enumerate(items):
                self.tbl.setItem(r, c, item)

    def _add_column(self):
        row = self.tbl.rowCount()
        self.tbl.insertRow(row)
        defaults = ['new_col', 'VARCHAR(255)', 'YES', '', '']
        for c, val in enumerate(defaults):
            self.tbl.setItem(row, c, QTableWidgetItem(val))

    def _delete_column(self):
        for r in set(i.row() for i in self.tbl.selectedItems()):
            col_name = self.tbl.item(r, 0).text() if self.tbl.item(r, 0) else ''
            reply = QMessageBox.question(self, '确认', f'删除列 {col_name}？此操作不可逆！')
            if reply == QMessageBox.Yes:
                self.tbl.removeRow(r)

    def _save(self):
        # 收集当前列定义
        new_cols = []
        for r in range(self.tbl.rowCount()):
            name = self.tbl.item(r, 0).text() if self.tbl.item(r, 0) else ''
            dtype = self.tbl.item(r, 1).text() if self.tbl.item(r, 1) else ''
            nullable = (self.tbl.item(r, 2).text() if self.tbl.item(r, 2) else 'YES') == 'YES'
            default = self.tbl.item(r, 3).text() if self.tbl.item(r, 3) else ''
            key = self.tbl.item(r, 4).text() if self.tbl.item(r, 4) else ''
            if name:
                new_cols.append({'name': name, 'type': dtype, 'nullable': nullable,
                                 'default': default, 'key': key})

        # 生成 ALTER 语句
        orig_names = {c['name'] for c in self.orig_columns}
        new_names = {c['name'] for c in new_cols}
        added = [c for c in new_cols if c['name'] not in orig_names]
        removed = [c for c in self.orig_columns if c['name'] not in new_names]
        modified = [c for c in new_cols if c['name'] in orig_names]

        sqls = []
        for c in added:
            null = '' if c['nullable'] else ' NOT NULL'
            dflt = f" DEFAULT '{c['default']}'" if c['default'] else ''
            sqls.append(f"ALTER TABLE `{self.db_name}`.`{self.table_name}` ADD COLUMN `{c['name']}` {c['type']}{null}{dflt}")

        for c in removed:
            sqls.append(f"ALTER TABLE `{self.db_name}`.`{self.table_name}` DROP COLUMN `{c['name']}`")

        for c in modified:
            orig = next((o for o in self.orig_columns if o['name'] == c['name']), None)
            if orig and (c['type'] != orig.get('type', '') or
                         c['nullable'] != orig.get('nullable', True) or
                         str(c.get('default', '')) != str(orig.get('default', ''))):
                null = '' if c['nullable'] else ' NOT NULL'
                dflt = f" DEFAULT '{c['default']}'" if c['default'] else ''
                sqls.append(f"ALTER TABLE `{self.db_name}`.`{self.table_name}` MODIFY COLUMN `{c['name']}` {c['type']}{null}{dflt}")

        if not sqls:
            QMessageBox.information(self, '提示', '没有变更')
            return

        preview = '\n'.join(sqls[:5])
        if len(sqls) > 5:
            preview += f'\n... 共 {len(sqls)} 条 ALTER 语句'
        reply = QMessageBox.question(self, '确认执行', f'将执行:\n{preview}\n\n确定？')
        if reply != QMessageBox.Yes:
            return

        # 执行
        errors = []
        for s in sqls:
            try:
                r = requests.post(f'{API_BASE}/api/execute',
                                  json={'sql': s, 'read_only': False}, timeout=10).json()
                if not r.get('success'):
                    errors.append(f'{s}: {r.get("error")}')
            except Exception as e:
                errors.append(f'{s}: {e}')

        if errors:
            QMessageBox.warning(self, '部分失败', '\n'.join(errors[:5]))
        else:
            QMessageBox.information(self, '成功', f'已执行 {len(sqls)} 条 ALTER 语句')
            self.accept()


# ═══════════════════════════════════════════
# AI 设置弹窗
# ═══════════════════════════════════════════
ENV_FILE = Path(__file__).parent.parent.parent / '.env'


class SettingsDialog(QDialog):
    """AI 密钥/模型设置"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('AI 设置')
        self.setMinimumWidth(480)
        self._build_ui()
        self._load_env()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        layout.addWidget(QLabel('<b>LLM API 配置</b>'))
        layout.addWidget(QLabel('修改后重启后端生效'))

        self.fields = {}
        field_defs = [
            ('DEEPSEEK_API_KEY', 'DeepSeek 密钥', True),
            ('OPENAI_API_KEY', 'OpenAI 密钥 (备用)', True),
            ('LLM_BASE_URL', 'API 地址', False),
            ('LLM_MODEL', '模型名称', False),
        ]

        form = QFormLayout()
        for key, label, is_pwd in field_defs:
            edit = QLineEdit()
            if is_pwd:
                edit.setEchoMode(QLineEdit.Password)
                # 加个显示/隐藏按钮
                show_btn = QPushButton('👁')
                show_btn.setFixedSize(28, 28)
                show_btn.setCheckable(True)
                show_btn.toggled.connect(lambda checked, e=edit: e.setEchoMode(
                    QLineEdit.Normal if checked else QLineEdit.Password))
                row = QHBoxLayout()
                row.addWidget(edit, 1)
                row.addWidget(show_btn)
                form.addRow(label, row)
            else:
                form.addRow(label, edit)
            self.fields[key] = edit
        layout.addLayout(form)

        # 保存/取消
        btns = QHBoxLayout()
        btns.addStretch()
        save_btn = QPushButton('保存到 .env')
        save_btn.setProperty('accent', True)
        save_btn.clicked.connect(self._save_env)
        btns.addWidget(save_btn)
        btns.addWidget(QPushButton('取消', clicked=self.reject))
        layout.addLayout(btns)

    def _load_env(self):
        """从 .env 和环境变量读取当前值"""
        import os as _os
        env_vars = {}
        if ENV_FILE.exists():
            for line in ENV_FILE.read_text('utf-8').split('\n'):
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    k, v = line.split('=', 1)
                    env_vars[k.strip()] = v.strip().strip('"').strip("'")

        for key, edit in self.fields.items():
            val = _os.getenv(key, '') or env_vars.get(key, '')
            edit.setText(val)

    def _save_env(self):
        """保存到 .env 文件"""
        import os as _os
        # 读取现有内容
        lines = []
        if ENV_FILE.exists():
            lines = ENV_FILE.read_text('utf-8').split('\n')

        # 更新或追加每个字段
        updated = set()
        new_lines = []
        for line in lines:
            stripped = line.strip()
            if stripped and not stripped.startswith('#'):
                for key in self.fields:
                    if stripped.startswith(f'{key}=') or stripped.startswith(f'{key} ='):
                        new_lines.append(f'{key}={self.fields[key].text()}')
                        updated.add(key)
                        break
                else:
                    new_lines.append(line)
            else:
                new_lines.append(line)

        # 追加未更新的字段
        for key, edit in self.fields.items():
            if key not in updated:
                new_lines.append(f'{key}={edit.text()}')
                updated.add(key)

        ENV_FILE.write_text('\n'.join(new_lines), 'utf-8')
        QMessageBox.information(self, '已保存', '设置已保存到 .env 文件。\n请重启后端 (uvicorn) 使配置生效。')
        self.accept()


# ═══════════════════════════════════════════
# 圆角容器控件
# ═══════════════════════════════════════════
class _RoundedWidget(QWidget):
    """带抗锯齿圆角的容器"""
    def paintEvent(self, event):
        from PySide6.QtGui import QPainter, QPainterPath, QBrush
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        path = QPainterPath()
        path.addRoundedRect(self.rect(), 10, 10)
        p.fillPath(path, QBrush(QColor('#2B2B2B')))
        # 裁剪子控件到圆角区域
        p.setClipPath(path)

    def resizeEvent(self, event):
        self.update()
        super().resizeEvent(event)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('箭毒蛙')
        self.resize(1200, 780)
        self.setMinimumSize(900, 600)

        self.PAGE_SIZE = 100
        self._current_db: Optional[Dict] = None
        self._c_collapsed = False
        self._c_saved_width = 0

        # 无边框窗口 + 圆角 + 拖拽跟踪
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self._drag_pos = None
        self._setup_ui()
        self._load_data()

    def _edge_test(self, pos):
        """检测鼠标是否在窗口边缘 (4px 精确线检测)"""
        r = self.rect()
        m = 4
        l = 0 <= pos.x() < m
        ri = r.width() - m < pos.x() <= r.width()
        t = 0 <= pos.y() < m
        b = r.height() - m < pos.y() <= r.height()
        if t and l: return Qt.TopLeftCorner
        if t and ri: return Qt.TopRightCorner
        if b and l: return Qt.BottomLeftCorner
        if b and ri: return Qt.BottomRightCorner
        if t: return Qt.TopEdge
        if b: return Qt.BottomEdge
        if l: return Qt.LeftEdge
        if ri: return Qt.RightEdge
        return None

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            edge = self._edge_test(event.position().toPoint())
            if edge and self.windowHandle():
                self.windowHandle().startSystemResize(edge)
                return
            if event.position().y() < 34:
                self._drag_pos = event.globalPosition().toPoint()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_pos is not None:
            delta = event.globalPosition().toPoint() - self._drag_pos
            self.move(self.pos() + delta)
            self._drag_pos = event.globalPosition().toPoint()
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._drag_pos = None
        super().mouseReleaseEvent(event)

    # ── UI 构建 ────────────────────────────
    def _setup_ui(self):
        # ── 自定义标题栏 ──
        titlebar = QWidget()
        titlebar.setFixedHeight(34)
        titlebar.setStyleSheet(f'background: #2B2B2B; border-bottom: 1px solid rgba(255,255,255,0.06);')
        tb_layout = QHBoxLayout(titlebar)
        tb_layout.setContentsMargins(10, 0, 0, 0)
        tb_layout.setSpacing(0)

        logo = QLabel('箭毒蛙')
        logo.setStyleSheet('color: #86909C; font-size: 12px; font-weight: 600; background: transparent; border: none;')
        tb_layout.addWidget(logo)
        tb_layout.addStretch()

        # 窗口控制按钮
        ctrl_style = 'QPushButton { background: transparent; border: none; color: #86909C; font-size: 14px; padding: 0 12px; } QPushButton:hover { background: rgba(255,255,255,0.06); } QPushButton#btnClose:hover { background: #E05555; color: white; }'
        min_btn = QPushButton('─')
        min_btn.setObjectName('btnMin')
        min_btn.setStyleSheet(ctrl_style)
        min_btn.clicked.connect(self.showMinimized)
        tb_layout.addWidget(min_btn)

        max_btn = QPushButton('□')
        max_btn.setStyleSheet(ctrl_style)
        max_btn.clicked.connect(lambda: self.showMaximized() if not self.isMaximized() else self.showNormal())
        tb_layout.addWidget(max_btn)

        close_btn = QPushButton('✕')
        close_btn.setObjectName('btnClose')
        close_btn.setStyleSheet(ctrl_style)
        close_btn.clicked.connect(self.close)
        tb_layout.addWidget(close_btn)

        # ── 菜单栏 ──
        menubar = QMenuBar()
        menubar.setStyleSheet(f"""
            QMenuBar {{ background: {BG}; color: {MUTED}; border-bottom: 1px solid rgba(255,255,255,0.06); padding: 2px 0; }}
            QMenuBar::item {{ padding: 4px 12px; }}
            QMenuBar::item:selected {{ background: {ACCENT_COLOR}; color: white; }}
        """)

        file_menu = menubar.addMenu('文件')
        act_conn = QAction('编辑连接', self)
        act_conn.triggered.connect(self._add_db_dialog)
        file_menu.addAction(act_conn)
        file_menu.addSeparator()
        act_exit = QAction('退出', self)
        act_exit.triggered.connect(self.close)
        file_menu.addAction(act_exit)

        set_menu = menubar.addMenu('设置')
        act_ai = QAction('AI 配置', self)
        act_ai.triggered.connect(self._open_settings)
        set_menu.addAction(act_ai)

        switch_menu = menubar.addMenu('切换')
        home_icon = QIcon(str(ICONS_DIR / 'home.png'))
        act_home = QAction(home_icon, '首页', self)
        act_home.triggered.connect(lambda: self._switch_workspace('home'))
        switch_menu.addAction(act_home)
        act_mysql = QAction('🐬 MySQL', self)
        act_mysql.triggered.connect(lambda: self._switch_workspace('mysql'))
        switch_menu.addAction(act_mysql)
        act_redis = QAction('🔴 Redis', self)
        act_redis.triggered.connect(lambda: self._switch_workspace('redis'))
        switch_menu.addAction(act_redis)

        tools_menu = menubar.addMenu('工具')
        act_export = QAction('导出数据 (CSV/Excel)', self)
        act_export.triggered.connect(self._export_csv)
        tools_menu.addAction(act_export)
        act_import = QAction('导入数据 (CSV/Excel/JSON)', self)
        act_import.triggered.connect(self._import_data)
        tools_menu.addAction(act_import)
        tools_menu.addSeparator()
        act_clear_hist = QAction('清除历史记录', self)
        act_clear_hist.triggered.connect(self._clear_history)
        tools_menu.addAction(act_clear_hist)

        central = _RoundedWidget(self)
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(3, 3, 3, 3)
        root_layout.setSpacing(0)

        root_layout.addWidget(titlebar)
        root_layout.addWidget(menubar)

        # QStackedWidget: Page0=首页, Page1=MySQL, Page2=Redis
        self.stack = QStackedWidget()
        root_layout.addWidget(self.stack)

        # ── Page 0: 首页 ──
        home = QWidget()
        home_layout = QVBoxLayout(home)
        home_layout.setAlignment(Qt.AlignCenter)
        home_layout.setSpacing(20)

        title = QLabel('箭毒蛙')
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet('font-size: 28px; font-weight: bold; color: #4A88C7;')
        home_layout.addWidget(title)

        subtitle = QLabel('选择数据源类型开始')
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet(f'font-size: 14px; color: {MUTED}; margin-bottom: 20px;')
        home_layout.addWidget(subtitle)

        cards = QHBoxLayout()
        cards.setAlignment(Qt.AlignCenter)
        cards.setSpacing(30)

        # MySQL 卡片 (用 QFrame 代替 QPushButton 避免文字截断)
        mysql_card = QFrame()
        mysql_card.setFixedSize(220, 280)
        mysql_card.setCursor(Qt.CursorShape.PointingHandCursor)
        mysql_card.setStyleSheet(f"""
            QFrame {{ background: {BG}; border: 2px solid rgba(37,116,255,0.2); border-radius: 12px; }}
            QFrame:hover {{ border-color: #00BFA5; background: #3C3F41; }}
        """)
        mysql_card.mousePressEvent = lambda e: self._switch_workspace('mysql')
        ml = QVBoxLayout(mysql_card)
        ml.setAlignment(Qt.AlignCenter)
        ml.setSpacing(8)
        mi = QLabel('🐬'); mi.setAlignment(Qt.AlignCenter); mi.setStyleSheet('font-size: 56px; border:none;')
        ml.addWidget(mi)
        mt = QLabel('MySQL'); mt.setAlignment(Qt.AlignCenter)
        mt.setStyleSheet('font-size: 20px; font-weight: bold; color: #A9B7C6; border:none;')
        ml.addWidget(mt)
        md = QLabel('关系型数据库\n表结构 · SQL查询 · 数据编辑'); md.setAlignment(Qt.AlignCenter)
        md.setStyleSheet(f'font-size: 12px; color: {MUTED}; border:none;')
        ml.addWidget(md)
        cards.addWidget(mysql_card)

        # Redis 卡片
        redis_card = QFrame()
        redis_card.setFixedSize(220, 280)
        redis_card.setCursor(Qt.CursorShape.PointingHandCursor)
        redis_card.setStyleSheet(f"""
            QFrame {{ background: {BG}; border: 2px solid rgba(220,50,50,0.2); border-radius: 12px; }}
            QFrame:hover {{ border-color: #DC3232; background: #3C3F41; }}
        """)
        redis_card.mousePressEvent = lambda e: self._switch_workspace('redis')
        rl = QVBoxLayout(redis_card)
        rl.setAlignment(Qt.AlignCenter)
        rl.setSpacing(8)
        ri = QLabel('🔴'); ri.setAlignment(Qt.AlignCenter); ri.setStyleSheet('font-size: 56px; border:none;')
        rl.addWidget(ri)
        rt = QLabel('Redis'); rt.setAlignment(Qt.AlignCenter)
        rt.setStyleSheet('font-size: 20px; font-weight: bold; color: #A9B7C6; border:none;')
        rl.addWidget(rt)
        rd = QLabel('键值数据库\n键浏览 · 值查看 · 缓存管理'); rd.setAlignment(Qt.AlignCenter)
        rd.setStyleSheet(f'font-size: 12px; color: {MUTED}; border:none;')
        rl.addWidget(rd)
        cards.addWidget(redis_card)

        home_layout.addLayout(cards)

        self.stack.addWidget(home)  # index 0

        # ── Page 1: MySQL workspace ──
        mysql_ws = QWidget()
        mysql_ws_layout = QVBoxLayout(mysql_ws)
        mysql_ws_layout.setContentsMargins(0, 0, 0, 0)
        mysql_ws_layout.setSpacing(0)
        splitter = QSplitter(Qt.Horizontal)
        panel_a = QWidget(); panel_b = QWidget(); panel_c = QWidget()
        self.panel_c = panel_c
        splitter.addWidget(panel_a); splitter.addWidget(panel_b); splitter.addWidget(panel_c)
        splitter.setStretchFactor(0, 0); splitter.setStretchFactor(1, 55); splitter.setStretchFactor(2, 45)
        splitter.setHandleWidth(4); panel_a.setFixedWidth(200)
        self._build_panel_a(panel_a)
        self._build_panel_b(panel_b)
        self._build_panel_c(panel_c)
        mysql_ws_layout.addWidget(splitter, 1)
        self.stack.addWidget(mysql_ws)  # index 1

        # ── Page 2: Redis workspace ──
        redis_ws = QWidget()
        redis_ws_layout = QVBoxLayout(redis_ws)
        redis_ws_layout.setContentsMargins(0, 0, 0, 0)
        redis_ws_layout.setSpacing(0)

        redis_splitter = QSplitter(Qt.Horizontal)
        ra = QWidget(); rb = QWidget(); rc = QWidget()
        redis_splitter.addWidget(ra); redis_splitter.addWidget(rb); redis_splitter.addWidget(rc)
        redis_splitter.setStretchFactor(0, 0); redis_splitter.setStretchFactor(1, 50); redis_splitter.setStretchFactor(2, 50)
        redis_splitter.setHandleWidth(4); ra.setFixedWidth(250)
        self._build_redis_panel_a(ra)
        self._build_redis_panel_b(rb)
        self._build_redis_panel_c(rc)
        redis_ws_layout.addWidget(redis_splitter, 1)
        self.stack.addWidget(redis_ws)  # index 2

        self.stack.setCurrentIndex(0)  # 默认首页

        # 状态栏已移除

    # ═══ A栏: 数据库菜单 ═══
    def _build_panel_a(self, parent: QWidget):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(4, 8, 2, 4)
        layout.setSpacing(4)

        hdr = QHBoxLayout()
        hdr.addWidget(QLabel('数据库'))
        hdr.addStretch()
        btn_style = 'QPushButton { background: transparent; border: none; padding: 2px 6px; color: #86909C; } QPushButton:hover { background: rgba(255,255,255,0.06); border-radius: 2px; color: #A9B7C6; }'
        add_btn = QPushButton('+ 新增')
        add_btn.setFixedHeight(20)
        add_btn.setStyleSheet(btn_style)
        add_btn.clicked.connect(self._add_db_dialog)
        hdr.addWidget(add_btn)
        refresh_btn = QPushButton('刷新')
        refresh_btn.setFixedHeight(20)
        refresh_btn.setStyleSheet(btn_style)
        refresh_btn.clicked.connect(self._load_schema_tree)
        hdr.addWidget(refresh_btn)
        layout.addLayout(hdr)

        self.schema_tree = QTreeWidget()
        self.schema_tree.setHeaderHidden(True)
        self.schema_tree.setIndentation(14)
        self.schema_tree.setIconSize(QSize(22, 22))
        self.schema_tree.itemExpanded.connect(self._on_tree_expand)
        self.schema_tree.itemClicked.connect(self._on_tree_click)
        self.schema_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.schema_tree.customContextMenuRequested.connect(self._on_tree_menu)
        layout.addWidget(self.schema_tree, 1)

        self.tree_label = QLabel('')
        self.tree_label.setStyleSheet(f'color: {MUTED}; font-size: 10px;')
        layout.addWidget(self.tree_label)

        conn_row = QHBoxLayout()
        self.sidebar_combo = QComboBox()
        self.sidebar_combo.setMinimumHeight(24)
        self.sidebar_combo.currentIndexChanged.connect(self._on_sidebar_db_select)
        conn_row.addWidget(self.sidebar_combo, 1)
        del_icon = QIcon(str(ICONS_DIR / 'cancel.png'))
        del_btn = QPushButton(del_icon, '')
        del_btn.setFixedSize(22, 22)
        del_btn.setIconSize(QSize(16, 16))
        del_btn.setFlat(True)
        del_btn.setToolTip('删除当前连接')
        del_btn.setStyleSheet('QPushButton { background: transparent; border: none; } QPushButton:hover { background: rgba(255,255,255,0.1); border-radius: 2px; }')
        del_btn.clicked.connect(self._delete_db_config)
        conn_row.addWidget(del_btn)
        mgr_btn = QPushButton('⚙')
        mgr_btn.setFixedSize(22, 22)
        mgr_btn.setToolTip('管理连接')
        mgr_btn.clicked.connect(self._add_db_dialog)
        conn_row.addWidget(mgr_btn)
        layout.addLayout(conn_row)

    # ═══ B栏: 数据浏览器 (多Tab) ═══
    def _build_panel_b(self, parent: QWidget):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(2, 4, 4, 4)
        layout.setSpacing(2)

        # 顶部状态行
        top = QHBoxLayout()
        self.rb_type = QLabel('')
        self.rb_elapsed = QLabel('')
        self.rb_rows = QLabel('')
        for lbl in [self.rb_type, self.rb_elapsed, self.rb_rows]:
            lbl.setStyleSheet(f'color: {MUTED}; font-size: 11px;')
            top.addWidget(lbl)
        self.rb_status = QLabel('')
        top.addWidget(self.rb_status)
        top.addStretch()
        self.expand_btn = QPushButton('▶ 展开AI')
        self.expand_btn.setFixedHeight(20)
        self.expand_btn.clicked.connect(self._toggle_panel_c)
        self.expand_btn.hide()
        top.addWidget(self.expand_btn)
        layout.addLayout(top)

        # 可关闭的 Tab 容器
        self.data_tabs = QTabWidget()
        self.data_tabs.setTabsClosable(True)
        self.data_tabs.tabCloseRequested.connect(self._close_data_tab)
        self.data_tabs.currentChanged.connect(self._on_tab_changed)
        self.data_tabs.setMovable(True)
        self._tab_close_icon = QIcon(str(ICONS_DIR / 'cancel.png'))
        layout.addWidget(self.data_tabs, 1)

        # 编辑工具栏
        edit_bar = QHBoxLayout()
        save_icon = QIcon(str(ICONS_DIR / 'diskette.png'))
        self.save_btn = QPushButton(save_icon, '保存修改')
        self.save_btn.setFixedHeight(24)
        self.save_btn.setProperty('accent', True)
        self.save_btn.clicked.connect(self._save_edits)
        self.save_btn.setEnabled(False)
        edit_bar.addWidget(self.save_btn)
        self.undo_btn = QPushButton('↩ 撤销')
        self.undo_btn.setFixedHeight(24)
        self.undo_btn.clicked.connect(self._undo_edits)
        self.undo_btn.setEnabled(False)
        edit_bar.addWidget(self.undo_btn)
        edit_bar.addStretch()
        add_btn = QPushButton('+ 新增行')
        add_btn.setFixedHeight(24)
        add_btn.clicked.connect(self._add_row_dialog)
        edit_bar.addWidget(add_btn)
        del_btn = QPushButton('🗑 删除行')
        del_btn.setFixedHeight(24)
        del_btn.clicked.connect(self._delete_selected_row)
        edit_bar.addWidget(del_btn)
        layout.addLayout(edit_bar)

        # 编辑跟踪
        self._edits: Dict[str, str] = {}
        # 排序/筛选状态
        self._sort_col: str = ''
        self._sort_dir: str = ''  # 'ASC' or 'DESC'
        self._filter_col: str = ''
        self._filter_val: str = ''

        # 翻页
        pager = QHBoxLayout()
        self.page_label = QLabel('')
        self.page_label.setStyleSheet(f'color: {MUTED}; font-size: 11px;')
        pager.addWidget(self.page_label)
        pager.addStretch()
        pb1 = QPushButton('← 上一页')
        pb1.setFixedHeight(24)
        pb1.clicked.connect(self._prev_page)
        pager.addWidget(pb1)
        pb2 = QPushButton('下一页 →')
        pb2.setFixedHeight(24)
        pb2.clicked.connect(self._next_page)
        pager.addWidget(pb2)
        layout.addLayout(pager)

        # Tab 数据存储: {tab_index: {'title','rows','columns','page','sql'}}
        self._tab_data: Dict[int, Dict] = {}
        self._add_data_tab(' 欢迎', [], ['提示'], is_temp=True)
        tbl = self._current_table()
        if tbl:
            tbl.setRowCount(1)
            item = QTableWidgetItem(' 点击左侧表名 或 在右侧执行SQL')
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            tbl.setItem(0, 0, item)

    # ═══ C栏: AI助手 + 日志 ═══
    def _build_panel_c(self, parent: QWidget):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(4, 4, 8, 4)
        layout.setSpacing(6)

        # 折叠按钮
        toggle_row = QHBoxLayout()
        toggle_row.addStretch()
        self.collapse_btn = QPushButton('◀ 隐藏AI')
        self.collapse_btn.setFixedHeight(20)
        self.collapse_btn.clicked.connect(self._toggle_panel_c)
        toggle_row.addWidget(self.collapse_btn)
        layout.addLayout(toggle_row)

        # DB 配置
        db_bar = QHBoxLayout()
        self.db_combo = QComboBox()
        self.db_combo.setMinimumHeight(26)
        self.db_combo.currentIndexChanged.connect(self._on_db_select)
        db_bar.addWidget(self.db_combo, 1)
        self.db_status = QLabel('')
        self.db_status.setStyleSheet(f'color: {MUTED}; font-size: 11px;')
        db_bar.addWidget(self.db_status)
        layout.addLayout(db_bar)

        # 自然语言输入
        nl_group = QGroupBox('自然语言输入')
        nl_ly = QVBoxLayout(nl_group)
        self.input_text = QPlainTextEdit()
        self.input_text.setPlaceholderText('用中文描述想要查询/修改的数据...\n例: 查询本月订单总数')
        self.input_text.setMaximumHeight(100)
        nl_ly.addWidget(self.input_text)
        btn_row = QHBoxLayout()
        self.gen_btn = QPushButton('生成 SQL')
        self.gen_btn.setProperty('accent', True)
        self.gen_btn.clicked.connect(self._generate_sql)
        btn_row.addWidget(self.gen_btn)
        btn_row.addWidget(QPushButton('清空', clicked=lambda: self.input_text.clear()))
        btn_row.addStretch()
        nl_ly.addLayout(btn_row)
        layout.addWidget(nl_group)

        # SQL 预览
        sql_group = QGroupBox('SQL 预览')
        sql_ly = QVBoxLayout(sql_group)
        self.sql_text = QTextEdit()
        self.sql_text.setPlaceholderText('-- AI 生成的 SQL 将显示在这里')
        self.sql_text.setReadOnly(True)
        self.sql_text.setLineWrapMode(QTextEdit.NoWrap)
        font = QFont('Consolas', 10)
        font.setStyleHint(QFont.Monospace)
        self.sql_text.setFont(font)
        sql_ly.addWidget(self.sql_text)
        self.highlighter = SqlHighlighter(self.sql_text.document())

        # SQL 补全词表 (右键 → 插入SQL片段)
        self.sql_completion_words: List[str] = []

        exec_row = QHBoxLayout()
        self.exec_btn = QPushButton('执行 SQL')
        self.exec_btn.setProperty('accent', True)
        self.exec_btn.clicked.connect(self._execute_sql)
        exec_row.addWidget(self.exec_btn)
        exec_row.addWidget(QPushButton('导出 CSV', clicked=self._export_csv))
        self.tx_check = QCheckBox('开启事务')
        exec_row.addWidget(self.tx_check)
        exec_row.addStretch()
        exec_row.addWidget(QPushButton('复制 SQL', clicked=self._copy_sql))
        sql_ly.addLayout(exec_row)
        layout.addWidget(sql_group, 1)

        # 日志 + 历史 (tabs)
        self.tabs = QTabWidget()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont('Consolas', 10))
        self.tabs.addTab(self.log_text, '执行日志')
        self._show_log('AI 生成的 SQL 分析、执行日志将显示在这里')

        self.history_list = QListWidget()
        self.history_list.itemDoubleClicked.connect(self._on_history_click)
        # 思考过程
        self.think_text = QTextEdit()
        self.think_text.setReadOnly(True)
        self.think_text.setFont(QFont('Consolas', 10))
        self.tabs.addTab(self.think_text, '思考过程')
        self._show_think('AI 思考过程将在这里实时显示...')

        self.tabs.addTab(self.history_list, '历史记录')
        layout.addWidget(self.tabs)

    def _open_settings(self):
        SettingsDialog(self).exec()

    def _switch_workspace(self, mode: str):
        """切换工作区: home / mysql / redis"""
        if mode == 'home':
            self.stack.setCurrentIndex(0)
        elif mode == 'mysql':
            self.stack.setCurrentIndex(1)
            if not hasattr(self, '_mysql_loaded'):
                self._load_data()
                self._mysql_loaded = True
        elif mode == 'redis':
            self.stack.setCurrentIndex(2)
            if not hasattr(self, '_redis_loaded'):
                self._load_redis_keys()
                self._redis_loaded = True

    # ═══ Redis Panel A: 键浏览 ═══
    def _build_redis_panel_a(self, parent: QWidget):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(4, 8, 2, 4)
        layout.setSpacing(4)

        hdr = QHBoxLayout()
        hdr.addWidget(QLabel('🔴 Redis 键'))
        hdr.addStretch()
        refresh_btn = QPushButton('刷新')
        refresh_btn.setFixedHeight(22)
        refresh_btn.clicked.connect(self._load_redis_keys)
        hdr.addWidget(refresh_btn)
        layout.addLayout(hdr)

        self.redis_key_input = QLineEdit()
        self.redis_key_input.setPlaceholderText('键模式 (如 user:*, * 查全部)')
        self.redis_key_input.returnPressed.connect(self._load_redis_keys)
        layout.addWidget(self.redis_key_input)

        self.redis_tree = QTreeWidget()
        self.redis_tree.setHeaderHidden(True)
        self.redis_tree.setIndentation(12)
        self.redis_tree.itemClicked.connect(self._on_redis_key_click)
        layout.addWidget(self.redis_tree, 1)

    def _build_redis_panel_b(self, parent: QWidget):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(4, 8, 8, 4)
        layout.setSpacing(4)

        self.redis_key_label = QLabel('选择一个键查看值')
        self.redis_key_label.setStyleSheet('font-weight: bold; font-size: 13px;')
        layout.addWidget(self.redis_key_label)

        type_row = QHBoxLayout()
        self.redis_type_label = QLabel('')
        self.redis_type_label.setStyleSheet(f'color: {MUTED};')
        type_row.addWidget(self.redis_type_label)
        type_row.addStretch()
        layout.addLayout(type_row)

        self.redis_value_text = QTextEdit()
        self.redis_value_text.setReadOnly(True)
        self.redis_value_text.setFont(QFont('Consolas', 10))
        layout.addWidget(self.redis_value_text, 1)

        save_row = QHBoxLayout()
        save_icon = QIcon(str(ICONS_DIR / 'diskette.png'))
        self.redis_save_btn = QPushButton(save_icon, '保存修改')
        self.redis_save_btn.setProperty('accent', True)
        self.redis_save_btn.clicked.connect(self._save_redis_value)
        self.redis_save_btn.setEnabled(False)
        save_row.addWidget(self.redis_save_btn)
        save_row.addStretch()
        self.redis_delete_btn = QPushButton('删除键')
        self.redis_delete_btn.setProperty('danger', True)
        self.redis_delete_btn.clicked.connect(self._delete_redis_key)
        self.redis_delete_btn.setEnabled(False)
        save_row.addWidget(self.redis_delete_btn)
        layout.addLayout(save_row)

    def _load_redis_keys(self):
        """加载 Redis 键列表"""
        pattern = self.redis_key_input.text() or '*'
        self.redis_tree.clear()

        def do_fetch():
            try:
                r = requests.get(f'{API_BASE}/api/redis/keys?pattern={pattern}', timeout=5)
                return r.json()
            except Exception as e:
                return {'error': str(e)}

        def callback(data):
            keys = data.get('keys', [])
            # 构建层级树
            tree_map: Dict[str, QTreeWidgetItem] = {}
            for key in sorted(keys):
                parts = key.split(':')
                parent = self.redis_tree
                for i, part in enumerate(parts):
                    full_path = ':'.join(parts[:i+1])
                    if i == len(parts) - 1:
                        item = QTreeWidgetItem([part])
                        item.setData(0, Qt.UserRole + 1, 'key')
                        item.setData(0, Qt.UserRole + 2, key)
                        type_color = data.get('types', {}).get(key, '')
                        if type_color:
                            item.setForeground(0, QColor(type_color))
                        if full_path in tree_map:
                            tree_map[full_path].parent().addChild(item)
                        else:
                            self.redis_tree.addTopLevelItem(item)
                    else:
                        if full_path not in tree_map:
                            item = QTreeWidgetItem([part])
                            self.redis_tree.addTopLevelItem(item)
                            tree_map[full_path] = item
                        parent = tree_map[full_path]

        self._run_async(do_fetch, callback)

    def _on_redis_key_click(self, item: QTreeWidgetItem, _col: int):
        if item.data(0, Qt.UserRole + 1) != 'key':
            return
        key = item.data(0, Qt.UserRole + 2)
        self.redis_key_label.setText(f'键: {key}')
        self.redis_value_text.setReadOnly(True)
        self.redis_save_btn.setEnabled(False)
        self.redis_delete_btn.setEnabled(False)

        def do_fetch():
            try:
                r = requests.get(f'{API_BASE}/api/redis/key/{key}', timeout=5)
                return r.json()
            except Exception as e:
                return {'error': str(e)}

        def callback(data):
            self.redis_type_label.setText(f'类型: {data.get("type", "")}')
            value = data.get('value', '')
            self.redis_value_text.setPlainText(str(value))
            self.redis_value_text.setReadOnly(False)
            self.redis_save_btn.setEnabled(True)
            self.redis_delete_btn.setEnabled(True)

        self._run_async(do_fetch, callback)

    # ═══ Redis Panel C: AI 助手 ═══
    def _build_redis_panel_c(self, parent: QWidget):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(4, 8, 8, 4)
        layout.setSpacing(6)

        nl_group = QGroupBox('Redis 自然语言')
        nl_ly = QVBoxLayout(nl_group)
        self.redis_input = QPlainTextEdit()
        self.redis_input.setPlaceholderText('用中文描述 Redis 操作...\n例: 查看所有user:开头的键、设置缓存key过期时间')
        self.redis_input.setMaximumHeight(80)
        nl_ly.addWidget(self.redis_input)
        btn_row = QHBoxLayout()
        self.redis_gen_btn = QPushButton('生成命令')
        self.redis_gen_btn.setProperty('accent', True)
        self.redis_gen_btn.clicked.connect(self._redis_generate)
        btn_row.addWidget(self.redis_gen_btn)
        btn_row.addWidget(QPushButton('清空', clicked=lambda: self.redis_input.clear()))
        btn_row.addStretch()
        nl_ly.addLayout(btn_row)
        layout.addWidget(nl_group)

        cmd_group = QGroupBox('Redis 命令预览')
        cmd_ly = QVBoxLayout(cmd_group)
        self.redis_cmd_text = QTextEdit()
        self.redis_cmd_text.setPlaceholderText('-- AI 生成的 Redis 命令将显示在这里')
        self.redis_cmd_text.setReadOnly(True)
        self.redis_cmd_text.setFont(QFont('Consolas', 10))
        cmd_ly.addWidget(self.redis_cmd_text)
        exec_row = QHBoxLayout()
        self.redis_exec_btn = QPushButton('执行命令')
        self.redis_exec_btn.setProperty('accent', True)
        self.redis_exec_btn.clicked.connect(self._redis_execute)
        exec_row.addWidget(self.redis_exec_btn)
        exec_row.addStretch()
        cmd_ly.addLayout(exec_row)
        layout.addWidget(cmd_group, 1)

        self.redis_log = QTextEdit()
        self.redis_log.setReadOnly(True)
        self.redis_log.setMaximumHeight(120)
        self.redis_log.setFont(QFont('Consolas', 10))
        layout.addWidget(self.redis_log)

    def _redis_generate(self):
        question = self.redis_input.toPlainText().strip()
        if not question:
            return
        self.redis_gen_btn.setText('生成中...')
        self.redis_gen_btn.setEnabled(False)

        def do_fetch():
            try:
                r = requests.post(f'{API_BASE}/api/redis/query',
                                  json={'question': question}, timeout=60)
                return r.json()
            except Exception as e:
                return {'error': str(e)}

        def callback(resp):
            self.redis_gen_btn.setText('生成命令')
            self.redis_gen_btn.setEnabled(True)
            if resp.get('error'):
                self.redis_cmd_text.setPlainText(f'# 错误: {resp["error"]}')
            else:
                self.redis_cmd_text.setPlainText(resp.get('command', resp.get('answer', '')))
            self.redis_log.append(f'[Q] {question}\n[A] {resp.get("command", resp.get("answer", ""))}\n')

        self._run_async(do_fetch, callback)

    def _redis_execute(self):
        cmd_text = self.redis_cmd_text.toPlainText().strip()
        if not cmd_text or cmd_text.startswith('#'):
            return
        self.redis_exec_btn.setText('执行中...')
        self.redis_exec_btn.setEnabled(False)

        def do_fetch():
            try:
                r = requests.post(f'{API_BASE}/api/redis/execute',
                                  json={'command': cmd_text}, timeout=10)
                return r.json()
            except Exception as e:
                return {'error': str(e)}

        def callback(resp):
            self.redis_exec_btn.setText('执行命令')
            self.redis_exec_btn.setEnabled(True)
            if resp.get('ok'):
                result = resp.get('result', 'OK')
                self.redis_log.append(f'✓ {result}\n')
                # 提取读命令结果展示到 B栏
                self._show_redis_result(result, cmd_text)
                self._load_redis_keys()
            else:
                self.redis_log.append(f'✗ {resp.get("error")}\n')

        self._run_async(do_fetch, callback)

    def _show_redis_result(self, result: str, cmd: str):
        """解析 Redis 执行结果并展示到 B栏"""
        lines = result.strip().split('\n')
        for line in lines:
            if 'GET:' in line or 'HGET:' in line or 'HGETALL:' in line or 'LRANGE:' in line or 'SMEMBERS:' in line or 'ZRANGE:' in line:
                # 提取 "CMD: value" 中的 value 部分
                val = line.split(':', 1)[-1].strip()
                self.redis_value_text.setPlainText(val)
                # 提取 key 名更新标题
                parts = cmd.strip().split()
                if len(parts) >= 2:
                    self.redis_key_label.setText(f'键: {parts[1]} (AI查询)')
                break

    def _save_redis_value(self):
        key = self.redis_key_label.text().replace('键: ', '')
        val = self.redis_value_text.toPlainText()
        try:
            r = requests.post(f'{API_BASE}/api/redis/key/{key}',
                              json={'value': val}, timeout=5).json()
            if r.get('ok'):
                QMessageBox.information(self, '成功', '已保存')
            else:
                QMessageBox.warning(self, '错误', str(r.get('error', '')))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))

    def _delete_redis_key(self):
        key = self.redis_key_label.text().replace('键: ', '')
        if QMessageBox.question(self, '确认', f'删除键 {key}?'):
            try:
                r = requests.delete(f'{API_BASE}/api/redis/key/{key}', timeout=5).json()
                if r.get('ok'):
                    self.redis_value_text.clear()
                    self._load_redis_keys()
            except Exception as e:
                QMessageBox.warning(self, '错误', str(e))

    # ── 数据加载 ────────────────────────────
    def _load_data(self):
        load_db_configs()
        self._refresh_db_combo()
        self._refresh_history()
        self._load_schema_tree()

    def _refresh_db_combo(self):
        self.db_combo.clear()
        self.sidebar_combo.clear()
        for c in _db_configs:
            label = f"{c['name']} ({c['type']})"
            self.db_combo.addItem(label, c)
            self.sidebar_combo.addItem(label, c)
        if _db_configs:
            self._current_db = _db_configs[0]
            self.sidebar_combo.setCurrentIndex(0)

    def _on_db_select(self, idx):
        if 0 <= idx < len(_db_configs):
            self._current_db = _db_configs[idx]
            self._load_schema_tree()

    def _on_sidebar_db_select(self, idx):
        """侧边栏连接切换 — 同步到主面板"""
        if 0 <= idx < len(_db_configs):
            self._current_db = _db_configs[idx]
            self.db_combo.setCurrentIndex(idx)
            self._load_schema_tree()

    # ── Schema 树 ────────────────────────────
    def _load_schema_tree(self):
        """加载所有数据库 → 表 → 列 到树控件"""
        self.schema_tree.clear()
        self.tree_label.setText('加载中...')

        def do_fetch():
            try:
                r = requests.get(f'{API_BASE}/api/databases', timeout=10)
                return r.json()
            except Exception as e:
                return {'error': str(e)}

        def callback(data):
            if 'error' in data:
                self.tree_label.setText('加载失败')
                return

            dbs = data.get('databases', [])
            self.tree_label.setText(f'{len(dbs)} 个数据库')
            # 更新 SQL 自动补全词表
            self._update_sql_completions()

            for db_name in dbs:
                db_item = QTreeWidgetItem([db_name])
                db_icon = QIcon(str(ICONS_DIR / 'big_database.png'))
                db_item.setIcon(0, db_icon)
                db_item.setData(0, Qt.UserRole + 1, 'database')
                db_item.setData(0, Qt.UserRole + 2, db_name)
                # 占位符, 展开时加载
                QTreeWidgetItem(db_item, ['...'])
                self.schema_tree.addTopLevelItem(db_item)

        self._run_async(do_fetch, callback)

    def _on_tree_click(self, item: QTreeWidgetItem, _col: int):
        """点击表节点 → B栏加载该表数据"""
        if item.data(0, Qt.UserRole + 1) != 'table':
            return
        table_name = item.text(0)
        db_name = item.data(0, Qt.UserRole + 2)
        self._load_table_data(db_name, table_name)

    def _load_table_data(self, db: str, table: str):
        """加载表数据 → 新建/切换Tab, 应用排序/筛选"""
        title = f' {table}'
        for i in range(self.data_tabs.count()):
            if self.data_tabs.tabText(i) == title:
                self.data_tabs.setCurrentIndex(i)
                break

        # 构建 SQL (带排序/筛选)
        sql = f'SELECT * FROM `{db}`.`{table}`'
        if self._filter_col and self._filter_val:
            sql += f" WHERE `{self._filter_col}` = '{self._filter_val}'"
        if self._sort_col:
            sql += f" ORDER BY `{self._sort_col}` {self._sort_dir or 'ASC'}"
        sql += ' LIMIT 500'

        def do_fetch():
            try:
                r = requests.post(f'{API_BASE}/api/execute',
                                  json={'sql': sql, 'read_only': True}, timeout=30)
                return r.json()
            except Exception as e:
                return {'error': str(e), 'success': False}

        def callback(resp):
            self.rb_type.setText(f'表: {table}')
            self.rb_elapsed.setText('')
            self.rb_status.setText('')
            self.rb_status.setStyleSheet('')
            if resp.get('success') and resp.get('data'):
                data = resp['data']
                # 获取主键列
                pk_col = ''
                try:
                    schema_r = requests.get(f'{API_BASE}/api/schema/{table}?database={db}', timeout=5).json()
                    for col in schema_r.get('columns', []):
                        if col.get('key') == 'PRI':
                            pk_col = col['name']
                            break
                except Exception:
                    pass

                idx = self._add_data_tab(title, data['columns'], data['rows'],
                                         sql=f'SELECT * FROM `{db}`.`{table}`',
                                         db_name=db, pk_col=pk_col)
                # 获取外键信息
                try:
                    fk_r = requests.get(
                        f'{API_BASE}/api/schema/{table}?database={db}', timeout=5).json()
                    fk_map = {}
                    for ci, col in enumerate(fk_r.get('columns', [])):
                        if col.get('key') == 'MUL':
                            # 查询外键引用 (简化: 同名列推断)
                            ref_parts = col.get('name', '').rsplit('_', 1)
                            ref_table = ref_parts[0] if len(ref_parts) > 1 else ''
                            if ref_table:
                                fk_map[ci] = (db, ref_table, 'id')
                    self._tab_data[idx]['fk_map'] = fk_map
                    # FK列标蓝
                    tbl = self._current_table()
                    if tbl and fk_map:
                        for r in range(tbl.rowCount()):
                            for c in fk_map:
                                item = tbl.item(r, c)
                                if item:
                                    item.setForeground(QColor('#6CB6FF'))
                                    item.setToolTip(f'点击跳转到 {fk_map[c][1]}')
                except Exception:
                    pass
                self._render_tab_page(idx)
                self.rb_rows.setText(f'{data["row_count"]} 行')
            else:
                err = resp.get('error', '未知错误')
                self.rb_rows.setText('0 行')
                self.rb_status.setText(f'✗ {err}')
                self.rb_status.setStyleSheet(f'color: {DANGER_COLOR}; font-weight: bold;')
                self._show_log(f'✗ 加载表 {table} 失败\n{err}')

        self._run_async(do_fetch, callback)

    # ── 数据 Tab 管理 ──────────────────────
    def _add_data_tab(self, title: str, columns: List[str], rows_data: List[List[Any]] = None,
                      is_temp: bool = False, sql: str = '', db_name: str = '', pk_col: str = ''):
        """新建数据Tab"""
        table = QTableWidget()
        table.setAlternatingRowColors(True)
        table.horizontalHeader().setStretchLastSection(True)
        table.setColumnCount(len(columns))
        table.setHorizontalHeaderLabels(columns)
        # 允许编辑 + 表头排序 + FK跳转
        table.cellChanged.connect(self._on_cell_edited)
        table.cellClicked.connect(self._on_cell_clicked)
        table.horizontalHeader().sectionClicked.connect(self._on_header_click)
        table.setContextMenuPolicy(Qt.CustomContextMenu)
        table.customContextMenuRequested.connect(self._on_table_context_menu)

        idx = self.data_tabs.addTab(table, title)
        self.data_tabs.setCurrentIndex(idx)
        # 透明嵌入式的关闭图标
        close_btn = QPushButton(self._tab_close_icon, '')
        close_btn.setFixedSize(20, 20)
        close_btn.setIconSize(QSize(12, 12))
        close_btn.setFlat(True)
        close_btn.setStyleSheet('''
            QPushButton { background: transparent; border: none; padding: 0; }
            QPushButton:hover { background: rgba(255,255,255,0.1); border-radius: 2px; }
        ''')
        close_btn.clicked.connect(lambda: self._close_data_tab(idx))
        self.data_tabs.tabBar().setTabButton(idx, QTabBar.RightSide, close_btn)

        info = {'title': title, 'columns': columns, 'rows': rows_data or [],
                'page': 0, 'sql': sql, 'is_temp': is_temp,
                'db_name': db_name, 'table_name': title.replace('📋 ', '').replace('🔍 ', ''),
                'pk_col': pk_col, 'fk_map': {}}  # fk_map: col_index → (ref_db, ref_table, ref_col)
        self._tab_data[idx] = info
        if rows_data:
            self._render_tab_page(idx)
        # 更新编辑按钮状态
        self._update_edit_buttons()
        return idx

    def _close_data_tab(self, idx: int):
        """关闭Tab"""
        if self.data_tabs.count() <= 1:
            return
        # 重建 tab_data 映射
        old_keys = list(self._tab_data.keys())
        self.data_tabs.removeTab(idx)
        new_data = {}
        new_idx = 0
        for old_key in old_keys:
            if old_key == idx:
                continue
            new_data[new_idx] = self._tab_data[old_key]
            new_idx += 1
        self._tab_data = new_data

    def _current_table(self) -> QTableWidget:
        """获取当前Tab的表格"""
        w = self.data_tabs.currentWidget()
        return w if isinstance(w, QTableWidget) else None

    def _current_tab_info(self) -> Dict:
        """获取当前Tab信息"""
        return self._tab_data.get(self.data_tabs.currentIndex(),
                                  {'rows': [], 'page': 0, 'columns': [], 'title': '', 'sql': '', 'is_temp': False})

    # ── 单元格编辑 ─────────────────────────
    def _on_cell_edited(self, row: int, col: int):
        """单元格被编辑时记录修改"""
        tbl = self._current_table()
        if not tbl:
            return
        info = self._current_tab_info()
        rows = info.get('rows', [])
        if row >= len(rows) or col >= len(rows[row]):
            return
        # 原始值
        orig = rows[row][col]
        orig_str = '' if orig is None else str(orig)
        new_val = tbl.item(row, col).text() if tbl.item(row, col) else ''
        # 只记录实际变更
        if new_val != orig_str:
            key = f'{row}_{col}'
            self._edits[key] = new_val
        elif f'{row}_{col}' in self._edits:
            del self._edits[f'{row}_{col}']
        self._update_edit_buttons()

    # ── 排序 & 筛选 ────────────────────────
    def _on_cell_clicked(self, row: int, col: int):
        """点击单元格 — 检查外键跳转"""
        info = self._current_tab_info()
        fk_map = info.get('fk_map', {})
        if col in fk_map:
            tbl = self._current_table()
            if not tbl or not tbl.item(row, col):
                return
            val = tbl.item(row, col).text()
            ref_db, ref_table, ref_col = fk_map[col]
            self._load_table_data(ref_db, ref_table)
            # 应用筛选跳转到对应行
            self._filter_col = ref_col
            self._filter_val = val
            self._reload_current_tab()
            self.rb_type.setText(f'FK: {ref_table}.{ref_col} = {val}')

    def _on_header_click(self, col: int):
        """点击表头切换排序"""
        info = self._current_tab_info()
        cols = info.get('columns', [])
        if col >= len(cols):
            return
        col_name = cols[col]
        if self._sort_col == col_name:
            self._sort_dir = 'DESC' if self._sort_dir == 'ASC' else 'ASC'
        else:
            self._sort_col = col_name
            self._sort_dir = 'ASC'
        self._reload_current_tab()
        self.rb_type.setText(f'排序: {col_name} {self._sort_dir}')

    def _on_table_context_menu(self, pos):
        """表格右键菜单 — 列筛选"""
        tbl = self._current_table()
        if not tbl:
            return
        col = tbl.columnAt(pos.x())
        if col < 0:
            return
        info = self._current_tab_info()
        cols = info.get('columns', [])
        if col >= len(cols):
            return
        col_name = cols[col]

        menu = QMenu(self)
        act_filter = QAction(f'筛选 "{col_name}" = ...', self)
        act_filter.triggered.connect(lambda: self._filter_column(col_name))
        menu.addAction(act_filter)

        act_clear = QAction('清除筛选/排序', self)
        act_clear.triggered.connect(self._clear_filter_sort)
        menu.addAction(act_clear)

        menu.exec(tbl.mapToGlobal(pos))

    def _filter_column(self, col_name: str):
        """弹出筛选输入框"""
        val, ok = tk.simpledialog.askstring('列筛选', f'{col_name} = ?', parent=self)  # noqa
        # Use QInputDialog instead
        from PySide6.QtWidgets import QInputDialog
        val, ok = QInputDialog.getText(self, '列筛选', f'WHERE {col_name} =')
        if ok:
            self._filter_col = col_name
            self._filter_val = val
            self._reload_current_tab()
            self.rb_type.setText(f'筛选: {col_name} = {val}')

    def _clear_filter_sort(self):
        self._sort_col = ''
        self._sort_dir = ''
        self._filter_col = ''
        self._filter_val = ''
        self._reload_current_tab()
        self.rb_type.setText('')

    def _update_edit_buttons(self):
        has_edits = bool(self._edits)
        self.save_btn.setEnabled(has_edits)
        self.undo_btn.setEnabled(has_edits)
        if has_edits:
            self.save_btn.setText(f'保存 ({len(self._edits)})')
        else:
            self.save_btn.setText('保存修改')

    def _undo_edits(self):
        """撤销所有编辑"""
        for key in list(self._edits.keys()):
            row_str, col_str = key.split('_')
            r, c = int(row_str), int(col_str)
            info = self._current_tab_info()
            rows = info.get('rows', [])
            if r < len(rows) and c < len(rows[r]):
                tbl = self._current_table()
                if tbl and tbl.item(r, c):
                    orig = rows[r][c]
                    tbl.item(r, c).setText('' if orig is None else str(orig))
                    if orig is None:
                        tbl.item(r, c).setForeground(QColor(MUTED))
        self._edits.clear()
        self._update_edit_buttons()

    def _save_edits(self):
        """保存所有修改到数据库"""
        if not self._edits:
            return
        info = self._current_tab_info()
        db_name = info.get('db_name', '')
        table_name = info.get('table_name', '')
        pk_col = info.get('pk_col', '')
        rows = info.get('rows', [])

        if not db_name or not table_name:
            # 尝试从 tab 标题推断
            title = info.get('title', '')
            # 标题格式: " users" 但 db_name 应该在 _load_table_data 中传入
            QMessageBox.warning(self, '提示', '无法确定数据库名，请先刷新表数据')
            return

        if not pk_col:
            pk_col = info.get('columns', ['id'])[0]  # 默认第一列是主键

        saved = 0
        errors = []
        for key, new_val in self._edits.items():
            row_str, col_str = key.split('_')
            r, c = int(row_str), int(col_str)
            if r >= len(rows):
                continue
            col_name = info['columns'][c] if c < len(info.get('columns', [])) else f'col_{c}'
            pk_val = str(rows[r][info['columns'].index(pk_col)]) if pk_col in info.get('columns', []) else str(r)

            try:
                resp = requests.post(f'{API_BASE}/api/table/update', json={
                    'database': db_name, 'table': table_name,
                    'pk_column': pk_col, 'pk_value': pk_val,
                    'column': col_name, 'value': new_val,
                }, timeout=10).json()
                if resp.get('success'):
                    saved += 1
                    # 更新本地缓存
                    rows[r][c] = new_val
                else:
                    detail = resp.get('detail', '失败')
                    # 提取 MySQL 错误信息
                    if isinstance(detail, str) and '|' in detail:
                        detail = detail.split('\n')[-1].strip() if '\n' in detail else detail
                    errors.append(f'{col_name}={new_val}: {detail}')
            except Exception as e:
                errors.append(f'{col_name}: {e}')

        self._edits.clear()
        self._update_edit_buttons()
        if errors:
            self.rb_status.setText(f'✗ {len(errors)} 个失败')
            self.rb_status.setStyleSheet(f'color: {DANGER_COLOR}; font-weight: bold;')
            self._show_log(f'保存结果: {saved} 成功, {len(errors)} 失败\n' + '\n'.join(errors))
        else:
            self.rb_status.setText(f'✓ 已保存 {saved} 处修改')
            self.rb_status.setStyleSheet(f'color: {SUCCESS_COLOR}; font-weight: bold;')
            self._render_tab_page()

    def _add_row_dialog(self):
        """新增行弹窗"""
        info = self._current_tab_info()
        if info.get('is_temp'):
            QMessageBox.warning(self, '提示', '请先打开一个数据表')
            return
        cols = info.get('columns', [])
        if not cols:
            return

        dialog = QDialog(self)
        dialog.setWindowTitle('新增行')
        dialog.setMinimumWidth(360)
        layout = QFormLayout(dialog)
        entries = {}
        for col in cols:
            e = QLineEdit()
            entries[col] = e
            layout.addRow(col, e)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(lambda: self._do_insert(info, entries, dialog))
        btns.rejected.connect(dialog.reject)
        layout.addRow(btns)
        dialog.exec()

    def _do_insert(self, info: Dict, entries: Dict[str, QLineEdit], dialog: QDialog):
        values = {k: e.text() for k, e in entries.items()}
        db_name = info.get('db_name', '')
        table_name = info.get('table_name', '')
        if not db_name or not table_name:
            return
        try:
            resp = requests.post(f'{API_BASE}/api/table/insert', json={
                'database': db_name, 'table': table_name, 'values': values
            }, timeout=10).json()
            if resp.get('success'):
                dialog.accept()
                self._show_log(f'✓ 已插入 1 行到 {table_name}')
                # 刷新当前 tab
                self._reload_current_tab()
            else:
                QMessageBox.warning(self, '错误', resp.get('detail', '插入失败'))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))

    def _delete_selected_row(self):
        """删除选中的行"""
        tbl = self._current_table()
        if not tbl:
            return
        rows = set()
        for item in tbl.selectedItems():
            rows.add(item.row())
        if not rows:
            QMessageBox.warning(self, '提示', '请先点击要删除的行')
            return
        r = list(rows)[0]
        info = self._current_tab_info()
        db_name = info.get('db_name', '')
        table_name = info.get('table_name', '')
        pk_col = info.get('pk_col', info.get('columns', ['id'])[0])
        all_rows = info.get('rows', [])

        if not db_name or not table_name or r >= len(all_rows):
            return
        pk_idx = info['columns'].index(pk_col) if pk_col in info.get('columns', []) else 0
        pk_val = str(all_rows[r][pk_idx])

        if not QMessageBox.question(self, '确认删除',
                                     f'确定删除 {table_name} 中 {pk_col}={pk_val} 的行?'):
            return

        try:
            resp = requests.post(f'{API_BASE}/api/table/delete', json={
                'database': db_name, 'table': table_name,
                'pk_column': pk_col, 'pk_value': pk_val,
            }, timeout=10).json()
            if resp.get('success'):
                self._show_log(f'✓ 已删除 {table_name} 中 {pk_col}={pk_val} 的行')
                self._reload_current_tab()
            else:
                QMessageBox.warning(self, '错误', resp.get('detail', '删除失败'))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))

    def _import_data(self):
        """导入CSV/Excel/JSON到当前表"""
        info = self._current_tab_info()
        table_name = info.get('table_name', '')
        db_name = info.get('db_name', '')
        if info.get('is_temp') or not table_name:
            QMessageBox.warning(self, '提示', '请先在A栏点击一个表，打开数据Tab')
            return

        path, _ = QFileDialog.getOpenFileName(
            self, '导入数据', '',
            'All (*.csv *.xlsx *.json);;CSV (*.csv);;Excel (*.xlsx);;JSON (*.json)')
        if not path:
            return

        try:
            rows_to_import = []
            if path.endswith('.csv'):
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    header = next(reader, None)
                    if header:
                        rows_to_import = [dict(zip(header, row)) for row in reader if row]
            elif path.endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(path)
                ws = wb.active
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows_to_import = [
                    dict(zip(header, [cell.value for cell in row]))
                    for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row)
                ]
            elif path.endswith('.json'):
                data = json.loads(Path(path).read_text('utf-8'))
                rows_to_import = data if isinstance(data, list) else [data]

            if not rows_to_import:
                QMessageBox.warning(self, '提示', '文件中没有数据')
                return

            # 确认
            cols_preview = list(rows_to_import[0].keys())[:5]
            reply = QMessageBox.question(
                self, '确认导入',
                f'将导入 {len(rows_to_import)} 行到 {table_name}\n'
                f'列: {", ".join(cols_preview)}...\n\n确认？')
            if reply != QMessageBox.Yes:
                return

            # 批量 INSERT
            imported = 0
            errors = []
            for row in rows_to_import:
                try:
                    resp = requests.post(f'{API_BASE}/api/table/insert', json={
                        'database': db_name, 'table': table_name, 'values': row
                    }, timeout=10).json()
                    if resp.get('success'):
                        imported += 1
                    else:
                        errors.append(resp.get('detail', ''))
                except Exception as e:
                    errors.append(str(e))

            self._show_log(f'导入完成: {imported} 成功 / {len(errors)} 失败')
            if errors:
                self._show_log(f'错误:\n' + '\n'.join(errors[:5]))
            self._reload_current_tab()
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))

    def _reload_current_tab(self):
        """刷新当前标签页数据"""
        info = self._current_tab_info()
        db_name = info.get('db_name', '')
        table_name = info.get('table_name', '')
        if db_name and table_name:
            self._load_table_data(db_name, table_name)

    def _on_tab_changed(self, idx: int):
        info = self._tab_data.get(idx)
        if info and info.get('rows'):
            self.page_label.setText(
                f'第 {info["page"]+1}/{max(1, (len(info["rows"])+99)//100)} 页 · 共 {len(info["rows"])} 行')

    # ── 折叠 ────────────────────────────────
    def _toggle_panel_c(self):
        """折叠/展开 AI 面板 (C栏)"""
        if self._c_collapsed:
            self.panel_c.setVisible(True)
            self.collapse_btn.setText('◀ 隐藏AI')
            self.expand_btn.hide()
        else:
            self.panel_c.setVisible(False)
            self.collapse_btn.setText('▶')
            self.expand_btn.show()
        self._c_collapsed = not self._c_collapsed

    def _update_sql_completions(self):
        """收集所有表名/列名 + SQL关键字用于自动补全"""
        words = set()
        # 收集已展开的表/列名
        for i in range(self.schema_tree.topLevelItemCount()):
            db_item = self.schema_tree.topLevelItem(i)
            for j in range(db_item.childCount()):
                tbl_item = db_item.child(j)
                table_name = tbl_item.text(0)
                words.add(table_name)
                for k in range(tbl_item.childCount()):
                    col_item = tbl_item.child(k)
                    if col_item.text(0) != '...':
                        col_name = col_item.text(0).split(':')[0]
                        words.add(col_name)
        # SQL 关键字
        words.update(['SELECT', 'FROM', 'WHERE', 'AND', 'OR', 'NOT', 'IN',
                       'LIKE', 'BETWEEN', 'JOIN', 'LEFT', 'RIGHT', 'INNER', 'ON',
                       'GROUP BY', 'ORDER BY', 'HAVING', 'LIMIT', 'OFFSET',
                       'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'ALTER', 'DROP',
                       'COUNT', 'SUM', 'AVG', 'MAX', 'MIN', 'DISTINCT', 'AS',
                       'NULL', 'IS NULL', 'IS NOT NULL', 'DEFAULT', 'PRIMARY KEY',
                       'LIMIT 100', 'ORDER BY', 'DESC', 'ASC'])
        self.sql_completion_words = sorted(words)

    def _on_tree_expand(self, item: QTreeWidgetItem):
        """展开节点时懒加载"""
        node_type = item.data(0, Qt.UserRole + 1)

        if node_type == 'database':
            # 加载数据库下的表
            item.takeChildren()
            db_name = item.data(0, Qt.UserRole + 2)

            def do_fetch():
                try:
                    r = requests.get(f'{API_BASE}/api/schema?database={db_name}', timeout=10)
                    return r.json()
                except Exception as e:
                    return {'error': str(e)}

            def callback(data):
                if 'error' in data:
                    return
                for t in data.get('tables', []):
                    t_item = QTreeWidgetItem([t['table']])
                    t_item.setData(0, Qt.UserRole + 1, 'table')
                    t_item.setData(0, Qt.UserRole + 2, db_name)
                    t_item.setData(0, Qt.UserRole, t.get('columns', []))
                    QTreeWidgetItem(t_item, ['...'])  # 占位列
                    item.addChild(t_item)

            self._run_async(do_fetch, callback)

        elif node_type == 'table':
            # 加载表的列
            item.takeChildren()
            columns = item.data(0, Qt.UserRole) or []
            for col in columns:
                key_info = ''
                if col.get('key') == 'PRI': key_info = ' 🔑'
                elif col.get('key') == 'MUL': key_info = ' 🔗'
                null = '?' if col.get('nullable') else ''
                col_text = f"{col['name']}: {col['type']}{null}{key_info}"
                col_item = QTreeWidgetItem([col_text])
                col_item.setData(0, Qt.UserRole + 1, 'column')
                col_item.setForeground(0, QColor(MUTED))
                item.addChild(col_item)

    def _on_tree_menu(self, pos):
        """右键菜单"""
        item = self.schema_tree.itemAt(pos)
        if not item:
            return
        node_type = item.data(0, Qt.UserRole + 1)

        if node_type == 'table':
            table_name = item.text(0)
            db_name = item.data(0, Qt.UserRole + 2)
            columns = item.data(0, Qt.UserRole) or []

            menu = QMenu(self)
            act_struct = QAction('查看表结构', self)
            act_struct.triggered.connect(lambda t=table_name, c=columns: self._show_table_structure(t, c))
            menu.addAction(act_struct)

            act_select = QAction(f'SELECT * FROM {table_name}', self)
            act_select.triggered.connect(lambda t=table_name: self.sql_text.setPlainText(
                f'SELECT * FROM {t} LIMIT 100;'))
            menu.addAction(act_select)

            act_design = QAction('📐 设计表 (可视化)', self)
            act_design.triggered.connect(lambda t=table_name, d=db_name, c=columns:
                                         self._open_table_designer(d, t, c))
            menu.addAction(act_design)
            act_ddl = QAction('📄 查看DDL', self)
            act_ddl.triggered.connect(lambda t=table_name, d=db_name: self._show_table_ddl(d, t))
            menu.addAction(act_ddl)

            act_copy = QAction('复制表名', self)
            act_copy.triggered.connect(lambda t=table_name: QApplication.clipboard().setText(t))
            menu.addAction(act_copy)

            menu.exec(self.schema_tree.mapToGlobal(pos))

        elif node_type == 'database':
            db_name = item.data(0, Qt.UserRole + 2)
            menu = QMenu(self)
            act_use = QAction(f'复制库名', self)
            act_use.triggered.connect(lambda d=db_name: QApplication.clipboard().setText(d))
            menu.addAction(act_use)
            menu.exec(self.schema_tree.mapToGlobal(pos))

    def _open_table_designer(self, db: str, table: str, columns: list):
        """打开可视化表设计器"""
        dlg = TableDesignerDialog(db, table, columns, self)
        if dlg.exec() == QDialog.Accepted:
            # 刷新表数据
            self._reload_current_tab()
            self._load_schema_tree()

    def _show_table_ddl(self, db: str, table: str):
        """显示建表 DDL"""
        def do_fetch():
            try:
                r = requests.get(f'{API_BASE}/api/table/ddl?database={db}&table={table}', timeout=10)
                return r.json()
            except Exception as e:
                return {'error': str(e)}

        def callback(resp):
            if resp.get('ddl'):
                self._show_log(f'-- 建表 DDL: {db}.{table}\n{resp["ddl"]}')
                self.tabs.setCurrentIndex(0)
            else:
                self._show_log(f'✗ 获取DDL失败: {resp.get("detail", "未知错误")}')
                self.tabs.setCurrentIndex(0)

        self._run_async(do_fetch, callback)

    def _show_table_structure(self, table_name: str, columns: list):
        """右键查看表结构 → 新建Tab"""
        rows = [[c.get('name', ''), c.get('type', ''), c.get('key', ''),
                 'YES' if c.get('nullable') else 'NO', str(c.get('default', ''))] for c in columns]
        idx = self._add_data_tab(f'🔍 {table_name}', ['字段', '类型', '键', '可空', '默认值'], rows)
        self._render_tab_page(idx)
        self.rb_type.setText(f'结构: {table_name}')
        self.rb_elapsed.setText('')
        self.rb_rows.setText(f'{len(columns)} 字段')
        self.rb_status.setText('')
        self.rb_status.setStyleSheet('')

    # ── DB 操作 ─────────────────────────────
    def _delete_db_config(self):
        """删除当前选中的数据库连接"""
        if not _db_configs:
            return
        idx = self.sidebar_combo.currentIndex()
        if idx < 0 or idx >= len(_db_configs):
            return
        cfg = _db_configs[idx]
        reply = QMessageBox.question(self, '确认删除',
                                      f'确定删除连接 "{cfg["name"]}" 吗？')
        if reply != QMessageBox.Yes:
            return
        del _db_configs[idx]
        save_db_configs()
        self._refresh_db_combo()
        self._load_schema_tree()

    def _add_db_dialog(self):
        dlg = DbConfigDialog(self)
        if dlg.exec() == QDialog.Accepted:
            cfg = dlg.get_config()
            if not cfg['name']:
                QMessageBox.warning(self, '提示', '请输入连接名称')
                return
            _db_configs.append(cfg)
            save_db_configs()
            self._refresh_db_combo()

    def _test_db(self):
        if not self._current_db:
            QMessageBox.warning(self, '提示', '请先选择数据库连接')
            return
        self.db_status.setText('检测中...')
        self.db_status.setStyleSheet(f'color: {WARNING_COLOR}; font-size: 11px;')

        def do_test():
            try:
                r = requests.post(f'{API_BASE}/api/execute',
                                  json={'sql': 'SELECT 1', 'read_only': True}, timeout=5)
                return r.json().get('success', False)
            except Exception:
                return False

        def callback(ok):
            if ok:
                self.db_status.setText('✓ 连接成功')
                self.db_status.setStyleSheet(f'color: {SUCCESS_COLOR}; font-size: 11px;')
            else:
                self.db_status.setText('✗ 连接失败')
                self.db_status.setStyleSheet(f'color: {DANGER_COLOR}; font-size: 11px;')

        self._run_async(do_test, callback)

    # ── 生成 SQL ────────────────────────────
    def _generate_sql(self):
        question = self.input_text.toPlainText().strip()
        if not question:
            QMessageBox.warning(self, '提示', '请输入问题')
            return
        self.gen_btn.setText('生成中...')
        self.gen_btn.setEnabled(False)

        def do_generate():
            try:
                # 追加提示确保 AI 返回 SQL
                prompt = question + '\n\n请直接给出SQL语句，不要额外解释。'
                r = requests.post(f'{API_BASE}/api/query',
                                  json={'question': prompt}, timeout=120)
                return r.json()
            except Exception as e:
                return {'error': str(e)}

        def callback(resp):
            self.gen_btn.setText('生成 SQL')
            self.gen_btn.setEnabled(True)
            sql = resp.get('sql', '') or ''
            error = resp.get('error', '') or ''
            answer = resp.get('answer', '') or ''
            steps = resp.get('steps', []) or ''

            # 显示思考过程
            think_lines = [f'📝 问题: {question}', '─' * 40]
            if steps:
                for i, s in enumerate(steps, 1):
                    think_lines.append(f'🔧 步骤{i}: {s}')
            if answer:
                think_lines.append(f'💡 分析: {answer[:500]}')
            self._show_think('\n'.join(think_lines))

            if error:
                self.sql_text.setPlainText(f'-- 生成失败: {error}')
                self._show_log(f'✗ 生成失败\n{error}')
            elif sql and sql.strip():
                self.sql_text.setPlainText(sql)
                self._append_log(f'[生成SQL] {question}\n{sql}\n')
            else:
                self.sql_text.setPlainText('-- AI 未生成 SQL 语句，请查看「思考过程」Tab')
                self._show_log(f'🤖 AI 分析 (未生成SQL)\n{"─"*50}\n{answer}')

        self._run_async(do_generate, callback)

    # ── 执行 SQL ────────────────────────────
    def _execute_sql(self):
        sql = self.sql_text.toPlainText().strip()
        if not sql or sql.startswith('--'):
            return

        # 高危拦截
        upper = sql.upper()
        for kw in DANGER_KW:
            if kw.upper() in upper:
                if kw == 'DELETE FROM' and 'WHERE' in upper:
                    continue
                reply = QMessageBox.question(
                    self, '高危操作',
                    f'检测到高危语句: {kw}\n\n该操作不可逆，确认执行？',
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return

        self.exec_btn.setText('执行中...')
        self.exec_btn.setEnabled(False)
        start = datetime.now()

        def do_execute():
            try:
                is_write = bool(re.match(r'^\s*(INSERT|UPDATE|DELETE|CREATE|ALTER|DROP|TRUNCATE)', sql, re.I))
                r = requests.post(f'{API_BASE}/api/execute',
                                  json={'sql': sql, 'read_only': not is_write}, timeout=30)
                data = r.json()
                data['_elapsed'] = (datetime.now() - start).total_seconds() * 1000
                data['_is_write'] = is_write
                return data
            except Exception as e:
                return {'error': str(e), 'success': False,
                        '_elapsed': (datetime.now() - start).total_seconds() * 1000}

        def callback(resp):
            self.exec_btn.setText('执行 SQL')
            self.exec_btn.setEnabled(True)
            elapsed = resp.get('_elapsed', 0)
            is_write = resp.get('_is_write', False)

            sql_type = 'SELECT'
            if is_write:
                sql_type = sql.strip().split()[0].upper()
            self.rb_type.setText(f'类型: {sql_type}')
            self.rb_elapsed.setText(f'耗时: {elapsed:.0f}ms')

            if resp.get('success'):
                data = resp.get('data')
                row_count = data.get('row_count', 0) if data else 0
                self.rb_rows.setText(f'行数: {row_count}')
                self.rb_status.setText('✓ 执行成功')
                self.rb_status.setStyleSheet(f'color: {SUCCESS_COLOR}; font-weight: bold;')

                if data and data.get('columns'):
                    self._show_table(data)
                    self._show_log(f'✓ 执行成功 · {row_count} 行 · {elapsed:.0f}ms\n结果已展示在 B 栏数据表格中')
                else:
                    self._show_log(f'✓ 执行成功\n受影响行数: {row_count}\n耗时: {elapsed:.0f}ms')
            else:
                self.rb_rows.setText('行数: —')
                self.rb_status.setText('✗ 执行失败')
                self.rb_status.setStyleSheet(f'color: {DANGER_COLOR}; font-weight: bold;')
                error = resp.get('error', '未知错误')
                self._show_log(f'✗ 执行失败\n{error}\n耗时: {elapsed:.0f}ms')
                self.tabs.setCurrentIndex(0)  # 切到日志 tab 显示错误

            # 保存历史
            history = load_history()
            history.append({
                'sql': sql[:200],
                'success': resp.get('success', False),
                'elapsed': f'{elapsed:.0f}ms',
                'time': datetime.now().strftime('%m/%d %H:%M'),
            })
            save_history(history)
            self._refresh_history()

        self._run_async(do_execute, callback)

    # ── 表格展示 ────────────────────────────
    def _show_table(self, data: Dict):
        """执行SQL结果 → 新建 查询结果 Tab"""
        cols = data['columns']
        rows = data['rows']
        idx = self._add_data_tab(' 查询结果', cols, rows, sql='')
        self._render_tab_page(idx)

    def _render_tab_page(self, tab_idx: int = -1):
        """渲染指定Tab的当前页"""
        if tab_idx < 0:
            tab_idx = self.data_tabs.currentIndex()
        info = self._tab_data.get(tab_idx)
        if not info or info.get('is_temp'):
            return
        tbl = self._current_table()
        if not tbl:
            return
        tbl.setRowCount(0)
        start = info['page'] * self.PAGE_SIZE
        rows = info['rows']
        page_rows = rows[start:start + self.PAGE_SIZE]
        tbl.setRowCount(len(page_rows))
        for r, row in enumerate(page_rows):
            for c, val in enumerate(row):
                display = '' if val is None else str(val)
                item = QTableWidgetItem(display)
                if val is None:
                    item.setForeground(QColor(MUTED))
                    item.setToolTip('NULL — 双击输入值')
                # 允许双击编辑（像 Navicat 一样）
                tbl.setItem(r, c, item)
        total = len(rows)
        pages = max(1, (total + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        self.page_label.setText(f'第 {info["page"]+1}/{pages} 页 · 共 {total} 行')

    def _prev_page(self):
        info = self._current_tab_info()
        if info['page'] > 0:
            info['page'] -= 1
            self._render_tab_page()

    def _next_page(self):
        info = self._current_tab_info()
        total = len(info.get('rows', []))
        pages = max(1, (total + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        if info['page'] < pages - 1:
            info['page'] += 1
            self._render_tab_page()

    # ── 日志 ────────────────────────────────
    def _append_log(self, msg: str):
        self.log_text.append(msg)

    def _show_think(self, msg: str):
        """显示/追加思考过程"""
        self.think_text.setPlainText(msg)
        self.tabs.setCurrentIndex(1)  # 思考过程 Tab

    def _append_think(self, msg: str):
        """追加思考内容(流式)"""
        self.think_text.append(msg)
        self.tabs.setCurrentIndex(1)

    def _show_log(self, msg: str):
        self.log_text.setPlainText(msg)

    # ── 历史记录 ────────────────────────────
    def _refresh_history(self):
        self.history_list.clear()
        history = load_history()
        for h in reversed(history[-50:]):
            status = '✓' if h.get('success') else '✗'
            label = f"{status} {h.get('time','')} | {h.get('elapsed','')} | {h.get('sql','')[:50]}"
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, h)
            self.history_list.addItem(item)
        # 历史记录已更新

    def _on_history_click(self, item: QListWidgetItem):
        h = item.data(Qt.UserRole)
        if h and h.get('sql'):
            self.sql_text.setPlainText(h['sql'])

    def _clear_history(self):
        reply = QMessageBox.question(self, '确认', '确定清除所有历史记录？')
        if reply == QMessageBox.Yes:
            save_history([])
            self._refresh_history()

    # ── 其他操作 ────────────────────────────
    def _copy_sql(self):
        sql = self.sql_text.toPlainText().strip()
        if sql and not sql.startswith('-- AI'):
            QApplication.clipboard().setText(sql)
            # SQL 已复制(无状态栏提示)

    def _export_csv(self):
        """导出当前 Tab 数据 (支持 CSV/Excel)"""
        info = self._current_tab_info()
        rows = info.get('rows', [])
        cols = info.get('columns', [])
        if not rows:
            QMessageBox.warning(self, '提示', '没有可导出的数据')
            return
        path, fmt = QFileDialog.getSaveFileName(
            self, '导出数据', '', 'Excel (*.xlsx);;CSV (*.csv)')
        if not path:
            return
        try:
            if path.endswith('.xlsx'):
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                wb = Workbook()
                ws = wb.active
                ws.title = info.get('title', 'Data')[:31]
                # 表头样式
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='2574FF', end_color='2574FF', fill_type='solid')
                for c, col_name in enumerate(cols, 1):
                    cell = ws.cell(row=1, column=c, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                for r, row in enumerate(rows, 2):
                    for c, val in enumerate(row, 1):
                        ws.cell(row=r, column=c, value=val)
                wb.save(path)
            else:
                with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                    w = csv.writer(f)
                    w.writerow(cols)
                    w.writerows(rows)
            QMessageBox.information(self, '提示', f'已导出到 {path}')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))

    # ── 工具 ────────────────────────────────
    def _run_async(self, target, callback):
        worker = ApiWorker(target)
        worker.finished.connect(callback)
        worker.start()
        # 保持引用防止 GC
        if not hasattr(self, '_workers'):
            self._workers = []
        self._workers.append(worker)


# ═══════════════════════════════════════════
# 入口
# ═══════════════════════════════════════════
def main():
    app = QApplication([])
    app.setStyleSheet(DARK_QSS)
    window = MainWindow()
    window.show()
    app.exec()


if __name__ == '__main__':
    main()
