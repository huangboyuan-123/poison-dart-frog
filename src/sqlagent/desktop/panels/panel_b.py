"""
Panel B: 数据浏览器 (多Tab管理)
"""
import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import requests
from PySide6.QtCore import Qt, QSize
from PySide6.QtGui import QColor, QIcon
from PySide6.QtWidgets import (
    QApplication, QDialog, QDialogButtonBox, QFileDialog,
    QFormLayout, QHBoxLayout, QLabel, QLineEdit, QMenu,
    QMessageBox, QPushButton, QTableWidget, QTableWidgetItem,
    QTabBar, QTabWidget, QVBoxLayout, QWidget, QInputDialog,
)

from ..constants import (
    API_BASE, BG, DANGER_COLOR, ICONS_DIR, MUTED, SUCCESS_COLOR,
)


class PanelBMixin:
    """Mixin providing Panel B (data browser, tab management) methods for MainWindow."""

    # ═══ B栏: 数据浏览器 (多Tab) ═══
    def _build_panel_b(self, parent: QWidget):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(2, 4, 4, 4)
        layout.setSpacing(2)

        # 顶部状态行
        top = QHBoxLayout()
        self.rb_type = QLabel('')
        self.rb_elapsed = QLabel('')
        self.rb_rows = QLabel('')
        for lbl in [self.rb_type, self.rb_elapsed, self.rb_rows]:
            lbl.setStyleSheet(f'color: {MUTED}; font-size: 11px;')
            top.addWidget(lbl)
        self.rb_status = QLabel('')
        top.addWidget(self.rb_status)
        top.addStretch()
        self.expand_btn = QPushButton('▶ 展开AI')
        self.expand_btn.setFixedHeight(20)
        self.expand_btn.clicked.connect(self._toggle_panel_c)
        self.expand_btn.hide()
        top.addWidget(self.expand_btn)
        layout.addLayout(top)

        # 可关闭的 Tab 容器
        self.data_tabs = QTabWidget()
        self.data_tabs.setTabsClosable(True)
        self.data_tabs.tabCloseRequested.connect(self._close_data_tab)
        self.data_tabs.currentChanged.connect(self._on_tab_changed)
        self.data_tabs.setMovable(True)
        self._tab_close_icon = QIcon(str(ICONS_DIR / 'cancel.png'))
        layout.addWidget(self.data_tabs, 1)

        # 编辑工具栏
        edit_bar = QHBoxLayout()
        save_icon = QIcon(str(ICONS_DIR / 'diskette.png'))
        self.save_btn = QPushButton(save_icon, '保存修改')
        self.save_btn.setFixedHeight(24)
        self.save_btn.setProperty('accent', True)
        self.save_btn.clicked.connect(self._save_edits)
        self.save_btn.setEnabled(False)
        edit_bar.addWidget(self.save_btn)
        self.undo_btn = QPushButton('↩ 撤销')  # undo emoji
        self.undo_btn.setFixedHeight(24)
        self.undo_btn.clicked.connect(self._undo_edits)
        self.undo_btn.setEnabled(False)
        edit_bar.addWidget(self.undo_btn)
        edit_bar.addStretch()
        add_btn = QPushButton('+ 新增行')
        add_btn.setFixedHeight(24)
        add_btn.clicked.connect(self._add_row_dialog)
        edit_bar.addWidget(add_btn)
        del_btn = QPushButton('\U0001f5d1 删除行')  # trash emoji
        del_btn.setFixedHeight(24)
        del_btn.clicked.connect(self._delete_selected_row)
        edit_bar.addWidget(del_btn)
        layout.addLayout(edit_bar)

        # 编辑跟踪
        self._edits: Dict[str, str] = {}
        self._sort_col: str = ''
        self._sort_dir: str = ''
        self._filter_col: str = ''
        self._filter_val: str = ''

        # 翻页
        pager = QHBoxLayout()
        self.page_label = QLabel('')
        self.page_label.setStyleSheet(f'color: {MUTED}; font-size: 11px;')
        pager.addWidget(self.page_label)
        pager.addStretch()
        pb1 = QPushButton('← 上一页')  # left arrow
        pb1.setFixedHeight(24)
        pb1.clicked.connect(self._prev_page)
        pager.addWidget(pb1)
        pb2 = QPushButton('下一页 →')  # right arrow
        pb2.setFixedHeight(24)
        pb2.clicked.connect(self._next_page)
        pager.addWidget(pb2)
        layout.addLayout(pager)

        # Tab 数据存储
        self._tab_data: Dict[int, Dict] = {}
        self._add_data_tab(' 欢迎', [], ['提示'], is_temp=True)
        tbl = self._current_table()
        if tbl:
            tbl.setRowCount(1)
            item = QTableWidgetItem(' 点击左侧表名 或 在右侧执行SQL')
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            tbl.setItem(0, 0, item)

    # ── 数据加载 ────────────────────────────
    def _load_table_data(self, db: str, table: str):
        """加载表数据 → 新建/切换Tab, 应用排序/筛选"""
        title = f' {table}'
        for i in range(self.data_tabs.count()):
            if self.data_tabs.tabText(i) == title:
                self.data_tabs.setCurrentIndex(i)
                break

        sql = f'SELECT * FROM `{db}`.`{table}`'
        if self._filter_col and self._filter_val:
            sql += f" WHERE `{self._filter_col}` = '{self._filter_val}'"
        if self._sort_col:
            sql += f" ORDER BY `{self._sort_col}` {self._sort_dir or 'ASC'}"
        sql += ' LIMIT 500'

        def do_fetch():
            try:
                r = requests.post(f'{API_BASE}/api/execute',
                                  json={'sql': sql, 'read_only': True}, timeout=30)
                return r.json()
            except Exception as e:
                return {'error': str(e), 'success': False}

        def callback(resp):
            self.rb_type.setText(f'表: {table}')
            self.rb_elapsed.setText('')
            self.rb_status.setText('')
            self.rb_status.setStyleSheet('')
            if resp.get('success') and resp.get('data'):
                data = resp['data']
                pk_col = ''
                try:
                    schema_r = requests.get(f'{API_BASE}/api/schema/{table}?database={db}', timeout=5).json()
                    for col in schema_r.get('columns', []):
                        if col.get('key') == 'PRI':
                            pk_col = col['name']
                            break
                except Exception:
                    pass

                idx = self._add_data_tab(title, data['columns'], data['rows'],
                                         sql=f'SELECT * FROM `{db}`.`{table}`',
                                         db_name=db, pk_col=pk_col)
                try:
                    fk_r = requests.get(
                        f'{API_BASE}/api/schema/{table}?database={db}', timeout=5).json()
                    fk_map = {}
                    for ci, col in enumerate(fk_r.get('columns', [])):
                        if col.get('key') == 'MUL':
                            ref_parts = col.get('name', '').rsplit('_', 1)
                            ref_table = ref_parts[0] if len(ref_parts) > 1 else ''
                            if ref_table:
                                fk_map[ci] = (db, ref_table, 'id')
                    self._tab_data[idx]['fk_map'] = fk_map
                    tbl = self._current_table()
                    if tbl and fk_map:
                        for r in range(tbl.rowCount()):
                            for c in fk_map:
                                item = tbl.item(r, c)
                                if item:
                                    item.setForeground(QColor('#6CB6FF'))
                                    item.setToolTip(f'点击跳转到 {fk_map[c][1]}')
                except Exception:
                    pass
                self._render_tab_page(idx)
                self.rb_rows.setText(f'{data["row_count"]} 行')
            else:
                err = resp.get('error', '未知错误')
                self.rb_rows.setText('0 行')
                self.rb_status.setText(f'✗ {err}')
                self.rb_status.setStyleSheet(f'color: {DANGER_COLOR}; font-weight: bold;')
                self._show_log(f'✗ 加载表 {table} 失败\n{err}')

        self._run_async(do_fetch, callback)

    # ── 数据 Tab 管理 ──────────────────────
    def _add_data_tab(self, title: str, columns: List[str], rows_data: List[List[Any]] = None,
                      is_temp: bool = False, sql: str = '', db_name: str = '', pk_col: str = ''):
        """新建数据Tab"""
        table = QTableWidget()
        # 统一暗色背景(不用系统交替行色)
        table.horizontalHeader().setStretchLastSection(True)
        table.setColumnCount(len(columns))
        table.setHorizontalHeaderLabels(columns)
        table.cellChanged.connect(self._on_cell_edited)
        table.cellClicked.connect(self._on_cell_clicked)
        table.horizontalHeader().sectionClicked.connect(self._on_header_click)
        table.setContextMenuPolicy(Qt.CustomContextMenu)
        table.customContextMenuRequested.connect(self._on_table_context_menu)

        idx = self.data_tabs.addTab(table, title)
        self.data_tabs.setCurrentIndex(idx)
        close_btn = QPushButton(self._tab_close_icon, '')
        close_btn.setFixedSize(20, 20)
        close_btn.setIconSize(QSize(12, 12))
        close_btn.setFlat(True)
        close_btn.setStyleSheet('''
            QPushButton { background: transparent; border: none; padding: 0; }
            QPushButton:hover { background: rgba(255,255,255,0.1); border-radius: 2px; }
        ''')
        close_btn.clicked.connect(lambda: self._close_data_tab(idx))
        self.data_tabs.tabBar().setTabButton(idx, QTabBar.RightSide, close_btn)

        info = {'title': title, 'columns': columns, 'rows': rows_data or [],
                'page': 0, 'sql': sql, 'is_temp': is_temp,
                'db_name': db_name, 'table_name': title.replace('\U0001f4cb ', '').replace('\U0001f50d ', ''),
                'pk_col': pk_col, 'fk_map': {}}
        self._tab_data[idx] = info
        if rows_data:
            self._render_tab_page(idx)
        self._update_edit_buttons()
        return idx

    def _close_data_tab(self, idx: int):
        """关闭Tab"""
        if self.data_tabs.count() <= 1:
            return
        old_keys = list(self._tab_data.keys())
        self.data_tabs.removeTab(idx)
        new_data = {}
        new_idx = 0
        for old_key in old_keys:
            if old_key == idx:
                continue
            new_data[new_idx] = self._tab_data[old_key]
            new_idx += 1
        self._tab_data = new_data

    def _current_table(self) -> QTableWidget:
        """获取当前Tab的表格"""
        w = self.data_tabs.currentWidget()
        return w if isinstance(w, QTableWidget) else None

    def _current_tab_info(self) -> Dict:
        """获取当前Tab信息"""
        return self._tab_data.get(self.data_tabs.currentIndex(),
                                  {'rows': [], 'page': 0, 'columns': [], 'title': '', 'sql': '', 'is_temp': False})

    def _on_tab_changed(self, idx: int):
        info = self._tab_data.get(idx)
        if info and info.get('rows'):
            self.page_label.setText(
                f'第 {info["page"]+1}/{max(1, (len(info["rows"])+99)//100)} 页 · 共 {len(info["rows"])} 行')

    # ── 渲染与翻页 ─────────────────────────
    def _render_tab_page(self, tab_idx: int = -1):
        """渲染指定Tab的当前页"""
        if tab_idx < 0:
            tab_idx = self.data_tabs.currentIndex()
        info = self._tab_data.get(tab_idx)
        if not info or info.get('is_temp'):
            return
        tbl = self._current_table()
        if not tbl:
            return
        tbl.setRowCount(0)
        start = info['page'] * self.PAGE_SIZE
        rows = info['rows']
        page_rows = rows[start:start + self.PAGE_SIZE]
        tbl.setRowCount(len(page_rows))
        for r, row in enumerate(page_rows):
            for c, val in enumerate(row):
                display = '' if val is None else str(val)
                item = QTableWidgetItem(display)
                if val is None:
                    item.setForeground(QColor(MUTED))
                    item.setToolTip('NULL — 双击输入值')
                tbl.setItem(r, c, item)
        total = len(rows)
        pages = max(1, (total + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        self.page_label.setText(f'第 {info["page"]+1}/{pages} 页 · 共 {total} 行')

    def _prev_page(self):
        info = self._current_tab_info()
        if info['page'] > 0:
            info['page'] -= 1
            self._render_tab_page()

    def _next_page(self):
        info = self._current_tab_info()
        total = len(info.get('rows', []))
        pages = max(1, (total + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        if info['page'] < pages - 1:
            info['page'] += 1
            self._render_tab_page()

    def _reload_current_tab(self):
        """刷新当前标签页数据"""
        info = self._current_tab_info()
        db_name = info.get('db_name', '')
        table_name = info.get('table_name', '')
        if db_name and table_name:
            self._load_table_data(db_name, table_name)

    # ── 单元格编辑 ─────────────────────────
    def _on_cell_edited(self, row: int, col: int):
        """单元格被编辑时记录修改"""
        tbl = self._current_table()
        if not tbl:
            return
        info = self._current_tab_info()
        rows = info.get('rows', [])
        if row >= len(rows) or col >= len(rows[row]):
            return
        orig = rows[row][col]
        orig_str = '' if orig is None else str(orig)
        new_val = tbl.item(row, col).text() if tbl.item(row, col) else ''
        if new_val != orig_str:
            key = f'{row}_{col}'
            self._edits[key] = new_val
        elif f'{row}_{col}' in self._edits:
            del self._edits[f'{row}_{col}']
        self._update_edit_buttons()

    def _update_edit_buttons(self):
        has_edits = bool(self._edits)
        self.save_btn.setEnabled(has_edits)
        self.undo_btn.setEnabled(has_edits)
        if has_edits:
            self.save_btn.setText(f'保存 ({len(self._edits)})')
        else:
            self.save_btn.setText('保存修改')

    def _undo_edits(self):
        """撤销所有编辑"""
        for key in list(self._edits.keys()):
            row_str, col_str = key.split('_')
            r, c = int(row_str), int(col_str)
            info = self._current_tab_info()
            rows = info.get('rows', [])
            if r < len(rows) and c < len(rows[r]):
                tbl = self._current_table()
                if tbl and tbl.item(r, c):
                    orig = rows[r][c]
                    tbl.item(r, c).setText('' if orig is None else str(orig))
                    if orig is None:
                        tbl.item(r, c).setForeground(QColor(MUTED))
        self._edits.clear()
        self._update_edit_buttons()

    def _save_edits(self):
        """保存所有修改到数据库"""
        if not self._edits:
            return
        info = self._current_tab_info()
        db_name = info.get('db_name', '')
        table_name = info.get('table_name', '')
        pk_col = info.get('pk_col', '')
        rows = info.get('rows', [])

        if not db_name or not table_name:
            QMessageBox.warning(self, '提示', '无法确定数据库名，请先刷新表数据')
            return

        if not pk_col:
            pk_col = info.get('columns', ['id'])[0]

        saved = 0
        errors = []
        for key, new_val in self._edits.items():
            row_str, col_str = key.split('_')
            r, c = int(row_str), int(col_str)
            if r >= len(rows):
                continue
            col_name = info['columns'][c] if c < len(info.get('columns', [])) else f'col_{c}'
            pk_val = str(rows[r][info['columns'].index(pk_col)]) if pk_col in info.get('columns', []) else str(r)

            try:
                resp = requests.post(f'{API_BASE}/api/table/update', json={
                    'database': db_name, 'table': table_name,
                    'pk_column': pk_col, 'pk_value': pk_val,
                    'column': col_name, 'value': new_val,
                }, timeout=10).json()
                if resp.get('success'):
                    saved += 1
                    rows[r][c] = new_val
                else:
                    detail = resp.get('detail', '失败')
                    if isinstance(detail, str) and '|' in detail:
                        detail = detail.split('\n')[-1].strip() if '\n' in detail else detail
                    errors.append(f'{col_name}={new_val}: {detail}')
            except Exception as e:
                errors.append(f'{col_name}: {e}')

        self._edits.clear()
        self._update_edit_buttons()
        if errors:
            self.rb_status.setText(f'✗ {len(errors)} 个失败')
            self.rb_status.setStyleSheet(f'color: {DANGER_COLOR}; font-weight: bold;')
            self._show_log(f'保存结果: {saved} 成功, {len(errors)} 失败\n' + '\n'.join(errors))
        else:
            self.rb_status.setText(f'✓ 已保存 {saved} 处修改')
            self.rb_status.setStyleSheet(f'color: {SUCCESS_COLOR}; font-weight: bold;')
            self._render_tab_page()

    # ── 排序 & 筛选 ────────────────────────
    def _on_cell_clicked(self, row: int, col: int):
        """点击单元格 — 检查外键跳转"""
        info = self._current_tab_info()
        fk_map = info.get('fk_map', {})
        if col in fk_map:
            tbl = self._current_table()
            if not tbl or not tbl.item(row, col):
                return
            val = tbl.item(row, col).text()
            ref_db, ref_table, ref_col = fk_map[col]
            self._load_table_data(ref_db, ref_table)
            self._filter_col = ref_col
            self._filter_val = val
            self._reload_current_tab()
            self.rb_type.setText(f'FK: {ref_table}.{ref_col} = {val}')

    def _on_header_click(self, col: int):
        """点击表头切换排序"""
        info = self._current_tab_info()
        cols = info.get('columns', [])
        if col >= len(cols):
            return
        col_name = cols[col]
        if self._sort_col == col_name:
            self._sort_dir = 'DESC' if self._sort_dir == 'ASC' else 'ASC'
        else:
            self._sort_col = col_name
            self._sort_dir = 'ASC'
        self._reload_current_tab()
        self.rb_type.setText(f'排序: {col_name} {self._sort_dir}')

    def _on_table_context_menu(self, pos):
        """表格右键菜单 — 列筛选"""
        tbl = self._current_table()
        if not tbl:
            return
        col = tbl.columnAt(pos.x())
        if col < 0:
            return
        info = self._current_tab_info()
        cols = info.get('columns', [])
        if col >= len(cols):
            return
        col_name = cols[col]

        menu = QMenu(self)
        # 选中单元格值
        item = tbl.itemAt(pos)
        cell_val = item.text() if item else ''

        act_filter = QAction(f'筛选 "{col_name}" = ...', self)
        act_filter.triggered.connect(lambda: self._filter_column(col_name))
        menu.addAction(act_filter)

        act_clear = QAction('清除筛选/排序', self)
        act_clear.triggered.connect(self._clear_filter_sort)
        menu.addAction(act_clear)

        # 格式化 JSON
        if cell_val and (cell_val.startswith('{') or cell_val.startswith('[')):
            menu.addSeparator()
            act_fmt = QAction('🎨 格式化 JSON', self)
            act_fmt.triggered.connect(lambda v=cell_val: self._show_json_formatted(v))
            menu.addAction(act_fmt)

        menu.exec(tbl.mapToGlobal(pos))

    def _filter_column(self, col_name: str):
        """弹出筛选输入框"""
        val, ok = QInputDialog.getText(self, '列筛选', f'WHERE {col_name} =')
        if ok:
            self._filter_col = col_name
            self._filter_val = val
            self._reload_current_tab()
            self.rb_type.setText(f'筛选: {col_name} = {val}')

    def _show_json_formatted(self, raw: str):
        """格式化JSON并彩色显示在D栏"""
        import json as _json, re as _re
        try:
            clean = _re.sub(r'(\d+)L\\b', r'\\1', raw)
            clean = clean.replace('None', 'null').replace('True', 'true').replace('False', 'false')
            data = _json.loads(clean)
            formatted = _json.dumps(data, indent=2, ensure_ascii=False)
            # 彩虹括号着色
            colors = ['#4FC3F7', '#66BB6A', '#FFEE58', '#AB47BC', '#EF5350', '#FFA726']
            html_lines = ['<pre style=\"font-family:Consolas;line-height:1.4;color:#A9B7C6;\">']
            for line in formatted.split('\\n'):
                depth = 0
                colored = []
                for ch in line:
                    if ch in '{[':
                        colored.append(f'<span style=\"color:{colors[depth % 6]};\"><b>{ch}</b></span>')
                        depth += 1
                    elif ch in '}]':
                        depth = max(0, depth - 1)
                        colored.append(f'<span style=\"color:{colors[depth % 6]};\"><b>{ch}</b></span>')
                    elif ch == '\"':
                        colored.append(f'<span style=\"color:#E8E8E8;\">{ch}</span>')
                    else:
                        colored.append(ch)
                html_lines.append(''.join(colored))
            html_lines.append('</pre>')
            self.think_text.setHtml(''.join(html_lines))
            self.tabs.setCurrentIndex(1)  # D栏
        except:
            QMessageBox.information(self, '提示', '无法解析为JSON')

    def _clear_filter_sort(self):
        self._sort_col = ''
        self._sort_dir = ''
        self._filter_col = ''
        self._filter_val = ''
        self._reload_current_tab()
        self.rb_type.setText('')

    # ── 行操作 ──────────────────────────────
    def _add_row_dialog(self):
        """新增行弹窗"""
        info = self._current_tab_info()
        if info.get('is_temp'):
            QMessageBox.warning(self, '提示', '请先打开一个数据表')
            return
        cols = info.get('columns', [])
        if not cols:
            return

        dialog = QDialog(self)
        dialog.setWindowTitle('新增行')
        dialog.setMinimumWidth(360)
        layout = QFormLayout(dialog)
        entries = {}
        for col in cols:
            e = QLineEdit()
            entries[col] = e
            layout.addRow(col, e)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(lambda: self._do_insert(info, entries, dialog))
        btns.rejected.connect(dialog.reject)
        layout.addRow(btns)
        dialog.exec()

    def _do_insert(self, info: Dict, entries: Dict[str, QLineEdit], dialog: QDialog):
        values = {k: e.text() for k, e in entries.items()}
        db_name = info.get('db_name', '')
        table_name = info.get('table_name', '')
        if not db_name or not table_name:
            return
        try:
            resp = requests.post(f'{API_BASE}/api/table/insert', json={
                'database': db_name, 'table': table_name, 'values': values
            }, timeout=10).json()
            if resp.get('success'):
                dialog.accept()
                self._show_log(f'✓ 已插入 1 行到 {table_name}')
                self._reload_current_tab()
            else:
                QMessageBox.warning(self, '错误', resp.get('detail', '插入失败'))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))

    def _delete_selected_row(self):
        """删除选中的行"""
        tbl = self._current_table()
        if not tbl:
            return
        rows = set()
        for item in tbl.selectedItems():
            rows.add(item.row())
        if not rows:
            QMessageBox.warning(self, '提示', '请先点击要删除的行')
            return
        r = list(rows)[0]
        info = self._current_tab_info()
        db_name = info.get('db_name', '')
        table_name = info.get('table_name', '')
        pk_col = info.get('pk_col', info.get('columns', ['id'])[0])
        all_rows = info.get('rows', [])

        if not db_name or not table_name or r >= len(all_rows):
            return
        pk_idx = info['columns'].index(pk_col) if pk_col in info.get('columns', []) else 0
        pk_val = str(all_rows[r][pk_idx])

        if not QMessageBox.question(self, '确认删除',
                                     f'确定删除 {table_name} 中 {pk_col}={pk_val} 的行?'):
            return

        try:
            resp = requests.post(f'{API_BASE}/api/table/delete', json={
                'database': db_name, 'table': table_name,
                'pk_column': pk_col, 'pk_value': pk_val,
            }, timeout=10).json()
            if resp.get('success'):
                self._show_log(f'✓ 已删除 {table_name} 中 {pk_col}={pk_val} 的行')
                self._reload_current_tab()
            else:
                QMessageBox.warning(self, '错误', resp.get('detail', '删除失败'))
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))

    def _import_data(self):
        """导入CSV/Excel/JSON到当前表"""
        info = self._current_tab_info()
        table_name = info.get('table_name', '')
        db_name = info.get('db_name', '')
        if info.get('is_temp') or not table_name:
            QMessageBox.warning(self, '提示', '请先在A栏点击一个表，打开数据Tab')
            return

        path, _ = QFileDialog.getOpenFileName(
            self, '导入数据', '',
            'All (*.csv *.xlsx *.json);;CSV (*.csv);;Excel (*.xlsx);;JSON (*.json)')
        if not path:
            return

        try:
            rows_to_import = []
            if path.endswith('.csv'):
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    header = next(reader, None)
                    if header:
                        rows_to_import = [dict(zip(header, row)) for row in reader if row]
            elif path.endswith('.xlsx'):
                from openpyxl import load_workbook
                wb = load_workbook(path)
                ws = wb.active
                header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows_to_import = [
                    dict(zip(header, [cell.value for cell in row]))
                    for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row)
                ]
            elif path.endswith('.json'):
                data = json.loads(Path(path).read_text('utf-8'))
                rows_to_import = data if isinstance(data, list) else [data]

            if not rows_to_import:
                QMessageBox.warning(self, '提示', '文件中没有数据')
                return

            cols_preview = list(rows_to_import[0].keys())[:5]
            reply = QMessageBox.question(
                self, '确认导入',
                f'将导入 {len(rows_to_import)} 行到 {table_name}\n'
                f'列: {", ".join(cols_preview)}...\n\n确认？')
            if reply != QMessageBox.Yes:
                return

            imported = 0
            errors = []
            for row in rows_to_import:
                try:
                    resp = requests.post(f'{API_BASE}/api/table/insert', json={
                        'database': db_name, 'table': table_name, 'values': row
                    }, timeout=10).json()
                    if resp.get('success'):
                        imported += 1
                    else:
                        errors.append(resp.get('detail', ''))
                except Exception as e:
                    errors.append(str(e))

            self._show_log(f'导入完成: {imported} 成功 / {len(errors)} 失败')
            if errors:
                self._show_log(f'错误:\n' + '\n'.join(errors[:5]))
            self._reload_current_tab()
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
