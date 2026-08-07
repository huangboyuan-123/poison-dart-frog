"""
Panel C: AI助手 + SQL执行 + 日志/历史
"""
import re
from datetime import datetime
from typing import List, Optional

import requests
from PySide6.QtCore import Qt, QSize
from PySide6.QtGui import QFont, QIcon, QAction
from PySide6.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QFileDialog, QGroupBox,
    QHBoxLayout, QLabel, QListWidget, QListWidgetItem,
    QMessageBox, QPlainTextEdit, QPushButton, QTabWidget,
    QTextEdit, QVBoxLayout, QWidget,
)

from ..constants import (
    API_BASE, BG, DANGER_COLOR, DANGER_KW, ICONS_DIR, MUTED,
    SUCCESS_COLOR, WARNING_COLOR, ACCENT_COLOR,
)
from ..highlighter import SqlHighlighter
from ..store import load_history, save_history
from ..utils import _extract_redis_commands, _extract_sql_from_stream, _md_to_html
from ..workers import StreamWorker, ApiWorker


class PanelCMixin:
    """Mixin providing Panel C (AI assistant) and Panel D (thinking) methods for MainWindow."""

    # ═══ C栏: AI助手 + 日志 ═══
    def _build_panel_c(self, parent: QWidget):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(4, 4, 8, 4)
        layout.setSpacing(6)

        # 折叠按钮
        toggle_row = QHBoxLayout()
        toggle_row.addStretch()
        self.collapse_btn = QPushButton('◀ 隐藏AI')  # left triangle
        self.collapse_btn.setFixedHeight(20)
        self.collapse_btn.clicked.connect(self._toggle_panel_c)
        toggle_row.addWidget(self.collapse_btn)
        think_toggle = QPushButton('思考')  # thought bubble
        think_toggle.setFixedHeight(20)
        think_toggle.clicked.connect(self._toggle_panel_d)
        toggle_row.addWidget(think_toggle)
        layout.addLayout(toggle_row)

        # DB 配置
        db_bar = QHBoxLayout()
        self.db_combo = QComboBox()
        self.db_combo.setMinimumHeight(26)
        self.db_combo.currentIndexChanged.connect(self._on_db_select)
        db_bar.addWidget(self.db_combo, 1)
        self.db_status = QLabel('')
        self.db_status.setStyleSheet(f'color: {MUTED}; font-size: 11px;')
        db_bar.addWidget(self.db_status)
        layout.addLayout(db_bar)

        # 自然语言输入
        nl_group = QGroupBox('自然语言输入')
        nl_ly = QVBoxLayout(nl_group)
        self.input_text = QPlainTextEdit()
        self.input_text.setPlaceholderText('用中文描述想要查询/修改的数据...\n例: 查询本月订单总数')
        self.input_text.setMaximumHeight(100)
        nl_ly.addWidget(self.input_text)
        btn_row = QHBoxLayout()
        self.gen_btn = QPushButton('生成 SQL')
        self.gen_btn.setProperty('accent', True)
        self.gen_btn.clicked.connect(self._generate_sql)
        btn_row.addWidget(self.gen_btn)
        btn_row.addWidget(QPushButton('清空', clicked=lambda: self.input_text.clear()))
        btn_row.addStretch()
        nl_ly.addLayout(btn_row)
        layout.addWidget(nl_group)

        # SQL 预览
        sql_group = QGroupBox('SQL 预览')
        sql_ly = QVBoxLayout(sql_group)
        self.sql_text = QTextEdit()
        self.sql_text.setPlaceholderText('-- AI 生成的 SQL 将显示在这里')
        self.sql_text.setReadOnly(False)  # 可手动写SQL
        self.sql_text.setLineWrapMode(QTextEdit.NoWrap)
        font = QFont('Consolas', 10)
        font.setStyleHint(QFont.Monospace)
        self.sql_text.setFont(font)
        sql_ly.addWidget(self.sql_text)
        self.highlighter = SqlHighlighter(self.sql_text.document())

        self.sql_completion_words: List[str] = []

        exec_row = QHBoxLayout()
        self.exec_btn = QPushButton('执行 SQL')
        self.exec_btn.setProperty('accent', True)
        self.exec_btn.clicked.connect(self._execute_sql)
        exec_row.addWidget(self.exec_btn)
        exec_row.addWidget(QPushButton('导出 CSV', clicked=self._export_csv))
        self.tx_check = QCheckBox('开启事务')
        exec_row.addWidget(self.tx_check)
        exec_row.addStretch()
        exec_row.addWidget(QPushButton('复制 SQL', clicked=self._copy_sql))
        sql_ly.addLayout(exec_row)
        layout.addWidget(sql_group, 1)

        # 日志 + 历史 (tabs)
        self.tabs = QTabWidget()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont('Consolas', 10))
        self.tabs.addTab(self.log_text, '执行日志')
        self._show_log('AI 生成的 SQL 分析、执行日志将显示在这里')

        self.history_list = QListWidget()
        self.history_list.itemDoubleClicked.connect(self._on_history_click)
        self.tabs.addTab(self.history_list, '历史记录')
        layout.addWidget(self.tabs)

    # ═══ D栏: AI思考过程 (可折叠) ═══
    def _build_panel_d(self, parent: QWidget):
        layout = QVBoxLayout(parent)
        layout.setContentsMargins(4, 8, 8, 4)
        layout.setSpacing(4)

        hdr = QHBoxLayout()
        hdr.addWidget(QLabel('思考过程'))  # thought bubble
        hdr.addStretch()
        self.d_close_btn = QPushButton('✕')  # X mark
        self.d_close_btn.setFixedSize(20, 20)
        self.d_close_btn.setFlat(True)
        self.d_close_btn.setStyleSheet('QPushButton { background: transparent; border: none; color: #86909C; } QPushButton:hover { color: #E05555; }')
        self.d_close_btn.clicked.connect(self._toggle_panel_d)
        hdr.addWidget(self.d_close_btn)
        layout.addLayout(hdr)

        self.think_text = QTextEdit()
        self.think_text.setReadOnly(True)
        self.think_text.setFont(QFont('Consolas', 10))
        layout.addWidget(self.think_text, 1)

    def _toggle_panel_d(self):
        """切换 D 栏显示/隐藏"""
        if self.panel_d.isVisible():
            self.panel_d.hide()
            self._splitter.setStretchFactor(3, 0)
        else:
            self.panel_d.show()
            self._splitter.setStretchFactor(3, 25)
            self._splitter.setStretchFactor(2, 30)

    def _toggle_panel_c(self):
        """折叠/展开 AI 面板 (C栏)"""
        if self._c_collapsed:
            self.panel_c.setVisible(True)
            self.collapse_btn.setText('◀ 隐藏AI')
            self.expand_btn.hide()
        else:
            self.panel_c.setVisible(False)
            self.collapse_btn.setText('▶')
            self.expand_btn.show()
        self._c_collapsed = not self._c_collapsed

    # ── 思考过程显示 ────────────────────────
    def _show_think(self, msg: str):
        """显示/追加思考过程"""
        self.think_text.setPlainText(msg)
        self.tabs.setCurrentIndex(1)

    def _append_think(self, msg: str):
        """追加思考内容(流式)"""
        self.think_text.append(msg)
        self.tabs.setCurrentIndex(1)

    # ── 日志 ────────────────────────────────
    def _show_log(self, msg: str):
        self.log_text.setPlainText(msg)

    def _append_log(self, msg: str):
        self.log_text.append(msg)

    # ── 生成 SQL ────────────────────────────
    def _generate_sql(self):
        question = self.input_text.toPlainText().strip()
        if not question:
            QMessageBox.warning(self, '提示', '请输入问题')
            return
        self.gen_btn.setText('生成中...')
        self.gen_btn.setEnabled(False)
        self._think_buffer = ''

        ws = getattr(self, '_current_workspace', 'mysql')
        if ws == 'redis':
            prompt = question
            stream_url = f'{API_BASE}/api/redis/query/stream'
        else:
            prompt = (question + '\n\n你是MySQL专家。步骤:1)用list_tables和get_table_schema查看结构 '
                      '2)根据结构生成SQL 3)**必须**输出```sql\n你的SQL语句;\n``` 不要只分析不写SQL！')
            stream_url = f'{API_BASE}/api/query/stream'

        self._stream_worker = StreamWorker(stream_url, {'question': prompt})

        full_buf = []
        def on_chunk(chunk):
            full_buf.append(chunk)
            self.think_text.setHtml(_md_to_html(''.join(full_buf)))
            self.think_text.verticalScrollBar().setValue(
                self.think_text.verticalScrollBar().maximum())
            if not self.panel_d.isVisible():
                self._toggle_panel_d()

        self._stream_worker.chunk.connect(on_chunk)

        def on_finished(resp):
            self.gen_btn.setText('生成 SQL' if ws != 'redis' else '生成命令')
            self.gen_btn.setEnabled(True)
            full = resp.get('full_text', '')
            error = resp.get('error', '')

            if error:
                self._show_think(f'❌ 错误: {error}')
                self.sql_text.setPlainText(f'-- 生成失败: {error}')
                return

            if ws == 'redis':
                cmds = _extract_redis_commands(full)
                self.sql_text.setPlainText('\n'.join(cmds) if cmds else full)
                self._append_log(f'[Redis] {question}\n{full[:300]}\n')
            else:
                sql = _extract_sql_from_stream(full)
                if sql:
                    self.sql_text.setPlainText(sql)
                    self._append_log(f'[生成SQL] {question}\n{sql}\n')
                else:
                    self.sql_text.setPlainText('-- AI 未生成 SQL，请查看思考过程')
                    self._append_log(f'[生成SQL] {question}\n{full[:300]}')

        self._stream_worker.finished.connect(on_finished)
        self._stream_worker.start()

    # ── 执行 SQL ────────────────────────────
    def _execute_sql(self):
        sql = self.sql_text.toPlainText().strip()
        if not sql or sql.startswith('--'):
            return

        ws = getattr(self, '_current_workspace', 'mysql')
        if ws == 'redis':
            self._execute_redis_cmd(sql)
            return

        # 高危拦截 (MySQL only)
        upper = sql.upper()
        for kw in DANGER_KW:
            if kw.upper() in upper:
                if kw == 'DELETE FROM' and 'WHERE' in upper:
                    continue
                reply = QMessageBox.question(
                    self, '高危操作',
                    f'检测到高危语句: {kw}\n\n该操作不可逆，确认执行？',
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply != QMessageBox.Yes:
                    return

        self.exec_btn.setText('执行中...')
        self.exec_btn.setEnabled(False)
        start = datetime.now()

        def do_execute():
            try:
                is_write = bool(re.match(r'^\s*(INSERT|UPDATE|DELETE|CREATE|ALTER|DROP|TRUNCATE)', sql, re.I))
                r = requests.post(f'{API_BASE}/api/execute',
                                  json={'sql': sql, 'read_only': not is_write}, timeout=30)
                data = r.json()
                data['_elapsed'] = (datetime.now() - start).total_seconds() * 1000
                data['_is_write'] = is_write
                return data
            except Exception as e:
                return {'error': str(e), 'success': False,
                        '_elapsed': (datetime.now() - start).total_seconds() * 1000}

        def callback(resp):
            self.exec_btn.setText('执行 SQL')
            self.exec_btn.setEnabled(True)
            elapsed = resp.get('_elapsed', 0)
            is_write = resp.get('_is_write', False)

            sql_type = 'SELECT'
            if is_write:
                sql_type = sql.strip().split()[0].upper()
            self.rb_type.setText(f'类型: {sql_type}')
            self.rb_elapsed.setText(f'耗时: {elapsed:.0f}ms')

            if resp.get('success'):
                data = resp.get('data')
                row_count = data.get('row_count', 0) if data else 0
                self.rb_rows.setText(f'行数: {row_count}')
                self.rb_status.setText('✓ 执行成功')
                self.rb_status.setStyleSheet(f'color: {SUCCESS_COLOR}; font-weight: bold;')

                if data and data.get('columns'):
                    self._show_table(data)
                    self._show_log(f'✓ 执行成功 · {row_count} 行 · {elapsed:.0f}ms\n结果已展示在 B 栏数据表格中')
                else:
                    self._show_log(f'✓ 执行成功\n受影响行数: {row_count}\n耗时: {elapsed:.0f}ms')
            else:
                self.rb_rows.setText('行数: —')  # em dash
                self.rb_status.setText('✗ 执行失败')
                self.rb_status.setStyleSheet(f'color: {DANGER_COLOR}; font-weight: bold;')
                error = resp.get('error', '未知错误')
                self._show_log(f'✗ 执行失败\n{error}\n耗时: {elapsed:.0f}ms')
                self.tabs.setCurrentIndex(0)

            history = load_history()
            history.append({
                'sql': sql[:200],
                'success': resp.get('success', False),
                'elapsed': f'{elapsed:.0f}ms',
                'time': datetime.now().strftime('%m/%d %H:%M'),
            })
            save_history(history)
            self._refresh_history()

        self._run_async(do_execute, callback)

    def _show_table(self, data: dict):
        """执行SQL结果 → 新建 查询结果 Tab"""
        cols = data['columns']
        rows = data['rows']
        idx = self._add_data_tab(' 查询结果', cols, rows, sql='')
        self._render_tab_page(idx)

    # ── 历史记录 ────────────────────────────
    def _refresh_history(self):
        self.history_list.clear()
        history = load_history()
        for h in reversed(history[-50:]):
            status = '✓' if h.get('success') else '✗'  # check / x
            label = f"{status} {h.get('time','')} | {h.get('elapsed','')} | {h.get('sql','')[:50]}"
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, h)
            self.history_list.addItem(item)

    def _on_history_click(self, item: QListWidgetItem):
        h = item.data(Qt.UserRole)
        if h and h.get('sql'):
            self.sql_text.setPlainText(h['sql'])

    def _clear_history(self):
        reply = QMessageBox.question(self, '确认', '确定清除所有历史记录？')
        if reply == QMessageBox.Yes:
            save_history([])
            self._refresh_history()

    # ── 其他操作 ────────────────────────────
    def _copy_sql(self):
        sql = self.sql_text.toPlainText().strip()
        if sql and not sql.startswith('-- AI'):
            QApplication.clipboard().setText(sql)

    def _export_csv(self):
        """导出当前 Tab 数据 (支持 CSV/Excel)"""
        info = self._current_tab_info()
        rows = info.get('rows', [])
        cols = info.get('columns', [])
        if not rows:
            QMessageBox.warning(self, '提示', '没有可导出的数据')
            return
        path, fmt = QFileDialog.getSaveFileName(
            self, '导出数据', '', 'Excel (*.xlsx);;CSV (*.csv)')
        if not path:
            return
        try:
            if path.endswith('.xlsx'):
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                wb = Workbook()
                ws = wb.active
                ws.title = info.get('title', 'Data')[:31]
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='2574FF', end_color='2574FF', fill_type='solid')
                for c, col_name in enumerate(cols, 1):
                    cell = ws.cell(row=1, column=c, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                for r, row in enumerate(rows, 2):
                    for c, val in enumerate(row, 1):
                        ws.cell(row=r, column=c, value=val)
                wb.save(path)
            else:
                import csv as csv_mod
                with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                    w = csv_mod.writer(f)
                    w.writerow(cols)
                    w.writerows(rows)
            QMessageBox.information(self, '提示', f'已导出到 {path}')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))

    # ── 工具 ────────────────────────────────
    def _run_async(self, target, callback):
        worker = ApiWorker(target)
        worker.finished.connect(callback)
        worker.start()
        if not hasattr(self, '_workers'):
            self._workers = []
        self._workers.append(worker)

    def _open_settings(self):
        from ..dialogs.settings import SettingsDialog
        SettingsDialog(self).exec()
